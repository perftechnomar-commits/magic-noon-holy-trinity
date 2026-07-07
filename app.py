from __future__ import annotations
from datetime import date, datetime, timedelta, timezone
from hashlib import sha256
from html import escape
from io import BytesIO
import gc
import hmac
import json
import os
import re
import time
from pathlib import Path
from typing import Any
from urllib.parse import urlencode, urljoin
from zoneinfo import ZoneInfo
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
import requests
import streamlit as st
from requests.auth import HTTPBasicAuth, HTTPDigestAuth
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# =============================================================================
# Configuration
# =============================================================================

APP_TITLE = "AtlasFlow"
APP_DIR = Path(__file__).resolve().parent
SNAPSHOT_DIR = APP_DIR / ".atlasflow_cache"
RAW_SNAPSHOT_FILE = SNAPSHOT_DIR / "reportdata_raw.parquet"
METADATA_SNAPSHOT_FILE = SNAPSHOT_DIR / "reportdata_metadata.json"
ODATA_ENDPOINT = "https://online.marorka.com/Odata/v1/ODataService.svc/ReportData"
REPORTPIVOTS_ENDPOINT = "https://online.marorka.com/Odata/v1/ODataService.svc/ReportPivots"
SHIPPIVOTS_ENDPOINT = "https://online.marorka.com/Odata/v1/ODataService.svc/ShipPivots"

SOURCE_CONFIGS = {
    "reportdata": {
        "label": "ReportData",
        "endpoint": ODATA_ENDPOINT,
        "snapshot_file": SNAPSHOT_DIR / "reportdata_raw.parquet",
        "metadata_file": SNAPSHOT_DIR / "reportdata_metadata.json",
        "datetime_candidates": ["StartDateTimeGMT", "DateTime", "dateTime", "Timestamp"],
    },
    "reportpivots": {
        "label": "ReportPivots",
        "endpoint": REPORTPIVOTS_ENDPOINT,
        "snapshot_file": SNAPSHOT_DIR / "reportpivots_raw.parquet",
        "metadata_file": SNAPSHOT_DIR / "reportpivots_metadata.json",
        "datetime_candidates": ["DateTime", "StartDateTimeGMT", "ReportDateTime", "Timestamp"],
    },
    "shippivots": {
        "label": "ShipPivots",
        "endpoint": SHIPPIVOTS_ENDPOINT,
        "snapshot_file": SNAPSHOT_DIR / "shippivots_raw.parquet",
        "metadata_file": SNAPSHOT_DIR / "shippivots_metadata.json",
        "datetime_candidates": ["DateTime", "StartDateTimeGMT", "Timestamp"],
    },
}
MAX_ODATA_PAGES = 250
API_CACHE_TTL_SECONDS = 21600  # 6 hours
API_FULL_START_DATE = date(2026, 1, 1)
TABLE_PREVIEW_ROW_LIMIT = 1000
DISPLAY_DATETIME_FORMAT = "%d/%m/%Y %H:%M"

EXCLUDED_REPORT_TYPES = [
    "Intake Report",
    "Fuel Change Report",
]

# ReportData is intentionally loaded in the same compact mode as the
# Performance KPIs app. Pulling every ValueDescription from ReportData is too
# broad for Streamlit Cloud and makes the source slow/fragile. These aliases
# are the KPI/calculation values that AtlasFlow needs now; we can expand this
# whitelist later in controlled groups.
PERFORMANCE_KPI_VALUE_ALIASES = {
    "Engine Distance [nm]": [
        "Engine Distance [nm]",
    ],
    "Distance Over Ground [nm]": [
        "Distance Over Ground [nm]",
    ],
    "Steaming Time Since Last Report [hh:mm]": [
        "Steaming Time Since Last Report [hh:mm]",
        "Steaming Time Since Last Report",
    ],
    "ME Load [%MCR]": [
        "ME Load [%MCR]",
        "ME Load [% MCR]",
    ],
    "Power from Torque Meter [kW]": [
        "Power from Torque Meter [kW]",
        "Total Shaft Power [kW] (kW)",
        "Total Shaft Power [kW]",
    ],
    "Main Engine - HSHFO": ["Main Engine - HSHFO"],
    "Main Engine - HSLFO": ["Main Engine - HSLFO"],
    "Main Engine - MGO": ["Main Engine - MGO"],
    "Main Engine - ULSHFO": ["Main Engine - ULSHFO"],
    "Main Engine - ULSLFO": ["Main Engine - ULSLFO"],
    "Main Engine - VLSHFO": ["Main Engine - VLSHFO"],
    "Main Engine - VLSLFO": ["Main Engine - VLSLFO"],
    "Boiler - HSHFO": ["Boiler - HSHFO"],
    "Boiler - HSLFO": ["Boiler - HSLFO"],
    "Boiler - MGO": ["Boiler - MGO"],
    "Boiler - ULSHFO": ["Boiler - ULSHFO"],
    "Boiler - ULSLFO": ["Boiler - ULSLFO"],
    "Boiler - VLSHFO": ["Boiler - VLSHFO"],
    "Boiler - VLSLFO": ["Boiler - VLSLFO"],

    # Additional bunker/fuel consumption ValueDescriptions for ME/DG/Auxiliary
    # analysis. These are included in the ReportData API whitelist so bunker
    # consumption fields can be selected/exported and later used for derived
    # calculations without broad-loading all ReportData.
    "Main Engine Total Consumed": [
        "Main Engine Total Consumed",
        "ME Total Consumed",
        "Main Engine Consumption",
        "ME Consumption",
        "MEConsumed",
    ],
    "Diesel Generator Total Consumed": [
        "Diesel Generator Total Consumed",
        "DG Total Consumed",
        "DG Totals Consumed",
        "DGTotalsConsumed",
        "DGTotalConsumed",
        "Generator Total Consumed",
    ],
    "Auxiliary Engine Total Consumed": [
        "Auxiliary Engine Total Consumed",
        "Aux Engine Total Consumed",
        "Aux Total Consumed",
        "AuxConsumed",
    ],
    "Total Fuel Consumed": [
        "Total Fuel Consumed",
        "Total Consumed",
        "Total Consumption",
        "Bunker Consumption",
        "Fuel Consumption",
    ],
    "Diesel Generator - HSHFO": ["Diesel Generator - HSHFO", "DG - HSHFO", "Generator - HSHFO"],
    "Diesel Generator - HSLFO": ["Diesel Generator - HSLFO", "DG - HSLFO", "Generator - HSLFO"],
    "Diesel Generator - MGO": ["Diesel Generator - MGO", "DG - MGO", "Generator - MGO", "DG - MGO/MDO"],
    "Diesel Generator - ULSHFO": ["Diesel Generator - ULSHFO", "DG - ULSHFO", "Generator - ULSHFO"],
    "Diesel Generator - ULSLFO": ["Diesel Generator - ULSLFO", "DG - ULSLFO", "Generator - ULSLFO"],
    "Diesel Generator - VLSHFO": ["Diesel Generator - VLSHFO", "DG - VLSHFO", "Generator - VLSHFO"],
    "Diesel Generator - VLSLFO": ["Diesel Generator - VLSLFO", "DG - VLSLFO", "Generator - VLSLFO"],
    "Auxiliary Engine - HSHFO": ["Auxiliary Engine - HSHFO", "Aux Engine - HSHFO", "Aux - HSHFO"],
    "Auxiliary Engine - HSLFO": ["Auxiliary Engine - HSLFO", "Aux Engine - HSLFO", "Aux - HSLFO"],
    "Auxiliary Engine - MGO": ["Auxiliary Engine - MGO", "Aux Engine - MGO", "Aux - MGO", "Aux - MGO/MDO"],
    "Auxiliary Engine - ULSHFO": ["Auxiliary Engine - ULSHFO", "Aux Engine - ULSHFO", "Aux - ULSHFO"],
    "Auxiliary Engine - ULSLFO": ["Auxiliary Engine - ULSLFO", "Aux Engine - ULSLFO", "Aux - ULSLFO"],
    "Auxiliary Engine - VLSHFO": ["Auxiliary Engine - VLSHFO", "Aux Engine - VLSHFO", "Aux - VLSHFO"],
    "Auxiliary Engine - VLSLFO": ["Auxiliary Engine - VLSLFO", "Aux Engine - VLSLFO", "Aux - VLSLFO"],

    # Lub oil ROB/received and DG running hours for consumption aggregation.
    # These are pulled from ReportData so AtlasFlow can expose oil consumption
    # totals as selectable derived variables, not as engineered KPI metrics.
    "MELO ROB [ltr]": ["MELO ROB [ltr]"],
    "MELO Received [ltr]": ["MELO Received [ltr]"],
    "Cylinder Oil 1 ROB [ltr]": ["Cylinder Oil 1 ROB [ltr]"],
    "Cylinder Oil 1 Received [ltr]": ["Cylinder Oil 1 Received [ltr]"],
    "Cylinder Oil 2 ROB [ltr]": ["Cylinder Oil 2 ROB [ltr]"],
    "Cylinder Oil 2 Received [ltr]": ["Cylinder Oil 2 Received [ltr]"],
    "GELO ROB [ltr]": ["GELO ROB [ltr]", "GELO Grade ROB [ltr]"],
    "GELO Received [ltr]": ["GELO Received [ltr]"],
    "DG1 Running Hours [hh:mm]": ["DG1 Running Hours [hh:mm]"],
    "DG2 Running Hours [hh:mm]": ["DG2 Running Hours [hh:mm]"],
    "DG3 Running Hours [hh:mm]": ["DG3 Running Hours [hh:mm]"],
    "DG4 Running Hours [hh:mm]": ["DG4 Running Hours [hh:mm]"],
}

REPORTDATA_VALUE_WHITELIST = sorted(
    {alias for aliases in PERFORMANCE_KPI_VALUE_ALIASES.values() for alias in aliases},
    key=str.casefold,
)
REPORTDATA_VALUE_WHITELIST_KEYS = {
    re.sub(r"[^a-z0-9]+", "", value.lower()) for value in REPORTDATA_VALUE_WHITELIST
}

SOURCE_COLUMNS = [
    "ReportId",
    "ShipName",
    "ReportType",
    "StartDateTimeGMT",
    "EndDateTimeGMT",
    "LapTime",
    "StateName",
    "ValueDescription",
    "ReportedValue",
]

PIVOT_IDENTITY_COLUMNS = [
    "ReportId",
    "ShipName",
    "ReportType",
    "StartDateTimeGMT",
    "EndDateTimeGMT",
    "LapTime",
    "StateName",
]

DEFAULT_DISPLAY_IDENTITY_COLUMNS = [
    "ShipName",
    "ReportType",
    "StartDateTimeGMT",
    "EndDateTimeGMT",
    "LapTime",
    "StateName",
]

# Derived calculation setup. These columns are calculated from the same Marorka
# source values used in the Performance KPIs app, but are exposed here as normal
# AtlasFlow variables that can be selected, displayed, filtered, summarized, and exported.
DERIVED_VALUE_ALIASES = {
    "Engine Distance [nm]": [
        "Engine Distance [nm]",
    ],
    "Distance Over Ground [nm]": [
        "Distance Over Ground [nm]",
    ],
    "Steaming Time Since Last Report [hh:mm]": [
        "Steaming Time Since Last Report [hh:mm]",
        "Steaming Time Since Last Report",
    ],
    "ME Load [%MCR]": [
        "ME Load [%MCR]",
        "ME Load [% MCR]",
    ],
    "Power from Torque Meter [kW]": [
        "Power from Torque Meter [kW]",
        "Total Shaft Power [kW] (kW)",
        "Total Shaft Power [kW]",
    ],
    "Main Engine - HSHFO": ["Main Engine - HSHFO"],
    "Main Engine - HSLFO": ["Main Engine - HSLFO"],
    "Main Engine - MGO": ["Main Engine - MGO"],
    "Main Engine - ULSHFO": ["Main Engine - ULSHFO"],
    "Main Engine - ULSLFO": ["Main Engine - ULSLFO"],
    "Main Engine - VLSHFO": ["Main Engine - VLSHFO"],
    "Main Engine - VLSLFO": ["Main Engine - VLSLFO"],
    "Boiler - HSHFO": ["Boiler - HSHFO"],
    "Boiler - HSLFO": ["Boiler - HSLFO"],
    "Boiler - MGO": ["Boiler - MGO"],
    "Boiler - ULSHFO": ["Boiler - ULSHFO"],
    "Boiler - ULSLFO": ["Boiler - ULSLFO"],
    "Boiler - VLSHFO": ["Boiler - VLSHFO"],
    "Boiler - VLSLFO": ["Boiler - VLSLFO"],

    # Additional bunker/fuel consumption ValueDescriptions for ME/DG/Auxiliary
    # analysis. These are included in the ReportData API whitelist so bunker
    # consumption fields can be selected/exported and later used for derived
    # calculations without broad-loading all ReportData.
    "Main Engine Total Consumed": [
        "Main Engine Total Consumed",
        "ME Total Consumed",
        "Main Engine Consumption",
        "ME Consumption",
        "MEConsumed",
    ],
    "Diesel Generator Total Consumed": [
        "Diesel Generator Total Consumed",
        "DG Total Consumed",
        "DG Totals Consumed",
        "DGTotalsConsumed",
        "DGTotalConsumed",
        "Generator Total Consumed",
    ],
    "Auxiliary Engine Total Consumed": [
        "Auxiliary Engine Total Consumed",
        "Aux Engine Total Consumed",
        "Aux Total Consumed",
        "AuxConsumed",
    ],
    "Total Fuel Consumed": [
        "Total Fuel Consumed",
        "Total Consumed",
        "Total Consumption",
        "Bunker Consumption",
        "Fuel Consumption",
    ],
    "Diesel Generator - HSHFO": ["Diesel Generator - HSHFO", "DG - HSHFO", "Generator - HSHFO"],
    "Diesel Generator - HSLFO": ["Diesel Generator - HSLFO", "DG - HSLFO", "Generator - HSLFO"],
    "Diesel Generator - MGO": ["Diesel Generator - MGO", "DG - MGO", "Generator - MGO", "DG - MGO/MDO"],
    "Diesel Generator - ULSHFO": ["Diesel Generator - ULSHFO", "DG - ULSHFO", "Generator - ULSHFO"],
    "Diesel Generator - ULSLFO": ["Diesel Generator - ULSLFO", "DG - ULSLFO", "Generator - ULSLFO"],
    "Diesel Generator - VLSHFO": ["Diesel Generator - VLSHFO", "DG - VLSHFO", "Generator - VLSHFO"],
    "Diesel Generator - VLSLFO": ["Diesel Generator - VLSLFO", "DG - VLSLFO", "Generator - VLSLFO"],
    "Auxiliary Engine - HSHFO": ["Auxiliary Engine - HSHFO", "Aux Engine - HSHFO", "Aux - HSHFO"],
    "Auxiliary Engine - HSLFO": ["Auxiliary Engine - HSLFO", "Aux Engine - HSLFO", "Aux - HSLFO"],
    "Auxiliary Engine - MGO": ["Auxiliary Engine - MGO", "Aux Engine - MGO", "Aux - MGO", "Aux - MGO/MDO"],
    "Auxiliary Engine - ULSHFO": ["Auxiliary Engine - ULSHFO", "Aux Engine - ULSHFO", "Aux - ULSHFO"],
    "Auxiliary Engine - ULSLFO": ["Auxiliary Engine - ULSLFO", "Aux Engine - ULSLFO", "Aux - ULSLFO"],
    "Auxiliary Engine - VLSHFO": ["Auxiliary Engine - VLSHFO", "Aux Engine - VLSHFO", "Aux - VLSHFO"],
    "Auxiliary Engine - VLSLFO": ["Auxiliary Engine - VLSLFO", "Aux Engine - VLSLFO", "Aux - VLSLFO"],

    # Lub oil ROB/received and DG running hours for consumption aggregation.
    # These are pulled from ReportData so AtlasFlow can expose oil consumption
    # totals as selectable derived variables, not as engineered KPI metrics.
    "MELO ROB [ltr]": ["MELO ROB [ltr]"],
    "MELO Received [ltr]": ["MELO Received [ltr]"],
    "Cylinder Oil 1 ROB [ltr]": ["Cylinder Oil 1 ROB [ltr]"],
    "Cylinder Oil 1 Received [ltr]": ["Cylinder Oil 1 Received [ltr]"],
    "Cylinder Oil 2 ROB [ltr]": ["Cylinder Oil 2 ROB [ltr]"],
    "Cylinder Oil 2 Received [ltr]": ["Cylinder Oil 2 Received [ltr]"],
    "GELO ROB [ltr]": ["GELO ROB [ltr]", "GELO Grade ROB [ltr]"],
    "GELO Received [ltr]": ["GELO Received [ltr]"],
    "DG1 Running Hours [hh:mm]": ["DG1 Running Hours [hh:mm]"],
    "DG2 Running Hours [hh:mm]": ["DG2 Running Hours [hh:mm]"],
    "DG3 Running Hours [hh:mm]": ["DG3 Running Hours [hh:mm]"],
    "DG4 Running Hours [hh:mm]": ["DG4 Running Hours [hh:mm]"],
}

ME_FUEL_COLUMNS = [
    "Main Engine - HSHFO",
    "Main Engine - HSLFO",
    "Main Engine - MGO",
    "Main Engine - ULSHFO",
    "Main Engine - ULSLFO",
    "Main Engine - VLSHFO",
    "Main Engine - VLSLFO",
]

BOILER_FUEL_COLUMNS = [
    "Boiler - HSHFO",
    "Boiler - HSLFO",
    "Boiler - MGO",
    "Boiler - ULSHFO",
    "Boiler - ULSLFO",
    "Boiler - VLSHFO",
    "Boiler - VLSLFO",
]

DG_FUEL_COLUMNS = [
    "Diesel Generator - HSHFO",
    "Diesel Generator - HSLFO",
    "Diesel Generator - MGO",
    "Diesel Generator - ULSHFO",
    "Diesel Generator - ULSLFO",
    "Diesel Generator - VLSHFO",
    "Diesel Generator - VLSLFO",
]

AUXILIARY_FUEL_COLUMNS = [
    "Auxiliary Engine - HSHFO",
    "Auxiliary Engine - HSLFO",
    "Auxiliary Engine - MGO",
    "Auxiliary Engine - ULSHFO",
    "Auxiliary Engine - ULSLFO",
    "Auxiliary Engine - VLSHFO",
    "Auxiliary Engine - VLSLFO",
]

DERIVED_VARIABLES = [
    "Calculated Slip",
    "ME Consumption Total",
    "DG Consumption Total",
    "Auxiliary Engine Consumption Total",
    "Boiler Sum",
    "Total Fuel Consumption",
    "Consumption ME 24 Hours [MT]",
    "SFOC [gr/Kwh]",
    "MELO Consumption Total [ltr]",
    "CYLO Consumption Total [ltr]",
    "GELO Consumption Total [ltr]",
    "Total DG Running Hours [hh:mm]",
]

VESSEL_GROUPS = {
    "Fleet 1": ["ATETI", "CMA CGM THALASSA", "CZECH", "DOLPHIN II", "GSL CHRISTEL ELISABETH", "GSL VINIA", "ORCA I", "MYNY", "SYDNEY EXPRESS"],
    "Fleet 2": ["AGIOS DIMITRIOS", "ELENI T", "MAIRA", "MELINA", "NEWYORKER", "NIKOLAS", "TORRANCE"],
    "Fleet 3": ["BREMERHAVEN EXPRESS", "CMA CGM ALCAZAR", "GSL ALICE", "GSL CHATEAU D'IF", "GSL ELEFTHERIA", "GSL MAREN", "GSL MELINA", "ISTANBUL EXPRESS"],
    "Fleet 4": ["ANTHEA Y", "COLOMBIA EXPRESS", "COSTA RICA EXPRESS", "JAMAICA EXPRESS", "MEXICO EXPRESS", "NICARAGUA EXPRESS", "PANAMA EXPRESS", "ZIM NORFOLK", "ZIM XIAMEN"],
    "Fleet 9": ["CMA CGM AMERICA", "CMA CGM SAMBHAR", "GSL ELENI", "GSL GRANIA", "GSL KALLIOPI", "GSL NINGBO", "MSC QINGDAO", "MSC TIANJIN"],
    "Fleet 10": ["CAPTAIN THANASIS I", "CMA CGM JAMAICA", "GSL CHRISTEN", "GSL NICOLETTA", "GSL VALERIE", "JULIE", "KUMASI", "MANET"],
    "Fleet 11": ["ATHENA", "EPAMINONDAS", "IAN H", "MARIANNA I", "MSC ROMA", "TINA I"],
    "Fleet 12": ["GSL DOROTHEA", "GSL KITHIRA", "GSL MARIA", "GSL MELITA", "GSL SYROS", "GSL TEGEA", "GSL TINOS", "GSL TRIPOLI"],
    "Fleet 14": ["GSL CHLOE", "GSL ELIZABETH", "GSL MAMITSA", "GSL MERCER", "GSL ROSSI", "GSL SUSAN", "TONSBERG"],
    "Fleet 15": ["GSL ALEXANDRA", "GSL ARCADIA", "GSL EFFIE", "GSL LYDIA", "GSL MYNY", "GSL SOFIA", "GSL VIOLETTA", "KOSTAS K", "MARIA Y"],
}

VESSEL_OPTIONS = sorted({v for vessels in VESSEL_GROUPS.values() for v in vessels})

st.set_page_config(page_title=APP_TITLE, layout="wide")


# =============================================================================
# Styling
# =============================================================================


def apply_custom_css() -> None:
    st.markdown(
        """
        <style>
        :root {
            --atlas-topbar-h: 67px;
            --atlas-sidebar-w: 324px;
            --atlas-ink: #0B1F33;
            --atlas-muted: #24364F;
            --atlas-soft: #64748B;
            --atlas-teal: #006B68;
            --atlas-teal-bright: #0AAEA6;
            --atlas-line: #D9E6E5;
            --atlas-bg: #FAFCFC;
            --atlas-chip: #DDF4F2;
        }

        html,
        body,
        .stApp {
            background: var(--atlas-bg) !important;
            color: var(--atlas-ink) !important;
            font-family: "Segoe UI", "Inter", "Aptos", Arial, sans-serif !important;
        }

        .stApp {
            background:
                linear-gradient(90deg, rgba(0, 107, 104, 0.035), transparent 24rem),
                linear-gradient(180deg, #FFFFFF 0%, #FBFDFD 46%, #F4FAF9 100%) !important;
        }

        header[data-testid="stHeader"] {
            left: 0 !important;
            right: 0 !important;
            width: 100vw !important;
            height: var(--atlas-topbar-h) !important;
            background: rgba(255, 255, 255, 0.98) !important;
            border-bottom: 1px solid rgba(15, 23, 42, 0.10) !important;
            box-shadow: 0 1px 12px rgba(15, 23, 42, 0.04) !important;
            z-index: 999990 !important;
        }

        header[data-testid="stHeader"] > div {
            height: var(--atlas-topbar-h) !important;
            background: transparent !important;
        }

        div[data-testid="stToolbar"] {
            top: 0.55rem !important;
            right: 1.55rem !important;
            z-index: 999995 !important;
        }

        div[data-testid="stDecoration"] {
            display: none !important;
        }

        .atlas-topbar-brand {
            position: fixed;
            top: 0;
            left: 0;
            height: var(--atlas-topbar-h);
            display: flex;
            align-items: center;
            gap: 0.95rem;
            padding-left: 1.35rem;
            z-index: 999996;
            pointer-events: none;
        }

        .atlas-menu-lines {
            width: 24px;
            height: 24px;
            position: relative;
            flex: 0 0 24px;
        }

        .atlas-menu-lines::before,
        .atlas-menu-lines::after,
        .atlas-menu-lines span {
            content: "";
            position: absolute;
            left: 2px;
            width: 18px;
            height: 2px;
            border-radius: 999px;
            background: var(--atlas-teal);
        }

        .atlas-menu-lines::before { top: 6px; }
        .atlas-menu-lines span { top: 11px; }
        .atlas-menu-lines::after { top: 16px; }

        .atlas-logo-mark {
            width: 34px;
            height: 34px;
            position: relative;
            flex: 0 0 34px;
        }

        .atlas-logo-mark::before,
        .atlas-logo-mark::after {
            content: "";
            position: absolute;
            border-radius: 18px 18px 18px 6px;
            transform: rotate(34deg);
            box-shadow: 0 4px 12px rgba(0, 107, 104, 0.18);
        }

        .atlas-logo-mark::before {
            width: 20px;
            height: 31px;
            left: 4px;
            top: 2px;
            background: linear-gradient(145deg, #013F43 0%, #008C86 62%, #19BFB5 100%);
        }

        .atlas-logo-mark::after {
            width: 18px;
            height: 24px;
            left: 15px;
            top: 10px;
            background: linear-gradient(145deg, #0FB5AD 0%, #006B68 100%);
            opacity: 0.94;
        }

        .atlas-brand-word {
            color: #07515A;
            font-size: 1.85rem;
            font-weight: 400;
            line-height: 1;
            letter-spacing: 0;
        }

        @media (min-width: 769px) {
            section[data-testid="stSidebar"] {
                width: var(--atlas-sidebar-w) !important;
                min-width: var(--atlas-sidebar-w) !important;
                top: var(--atlas-topbar-h) !important;
                height: calc(100vh - var(--atlas-topbar-h)) !important;
                background:
                    radial-gradient(circle at 18px 8px, rgba(35, 209, 199, 0.18), transparent 16rem),
                    linear-gradient(180deg, #006A66 0%, #004743 46%, #003C39 100%) !important;
                border-right: 1px solid rgba(255, 255, 255, 0.16) !important;
                box-shadow: none !important;
            }

            section[data-testid="stSidebar"] > div {
                height: calc(100vh - var(--atlas-topbar-h)) !important;
                padding: 1.25rem 1.25rem 6rem 1.25rem !important;
            }

            .block-container {
                max-width: none !important;
                padding: 3.55rem 1.9rem 3rem 2.35rem !important;
            }
        }

        @media (max-width: 768px) {
            .atlas-brand-word { font-size: 1.35rem; }
            .atlas-logo-mark { width: 28px; height: 28px; }
            .block-container { padding: 5.2rem 1rem 2rem 1rem !important; }
        }

        section[data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
            gap: 0.72rem !important;
        }

        section[data-testid="stSidebar"] h1,
        section[data-testid="stSidebar"] h2,
        section[data-testid="stSidebar"] h3,
        section[data-testid="stSidebar"] p,
        section[data-testid="stSidebar"] span,
        section[data-testid="stSidebar"] label,
        section[data-testid="stSidebar"] label *,
        section[data-testid="stSidebar"] [data-testid="stWidgetLabel"],
        section[data-testid="stSidebar"] [data-testid="stWidgetLabel"] *,
        section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] *,
        section[data-testid="stSidebar"] [data-testid="stCaptionContainer"] * {
            color: #FFFFFF !important;
            -webkit-text-fill-color: #FFFFFF !important;
            font-weight: 760 !important;
            letter-spacing: 0 !important;
        }

        section[data-testid="stSidebar"] h3 {
            font-size: 1.05rem !important;
            margin-top: 0.25rem !important;
            margin-bottom: 0.1rem !important;
        }

        section[data-testid="stSidebar"] [data-testid="stCaptionContainer"] *,
        section[data-testid="stSidebar"] small {
            color: #D7FFFA !important;
            -webkit-text-fill-color: #D7FFFA !important;
        }

        section[data-testid="stSidebar"] div[data-baseweb="select"] > div,
        section[data-testid="stSidebar"] div[data-baseweb="input"],
        section[data-testid="stSidebar"] [data-testid="stMultiSelect"] div[data-baseweb="select"] > div,
        section[data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"] > div {
            min-height: 44px !important;
            border-radius: 9px !important;
            border: 1px solid rgba(8, 72, 70, 0.12) !important;
            background: #FFFFFF !important;
            box-shadow: none !important;
        }

        section[data-testid="stSidebar"] div[data-baseweb="select"] > div *,
        section[data-testid="stSidebar"] div[data-baseweb="input"] *,
        section[data-testid="stSidebar"] input {
            color: var(--atlas-ink) !important;
            -webkit-text-fill-color: var(--atlas-ink) !important;
        }

        section[data-testid="stSidebar"] div[data-testid="stButton"] button,
        section[data-testid="stSidebar"] div[data-testid="stButton"] button *,
        section[data-testid="stSidebar"] .stButton button,
        section[data-testid="stSidebar"] .stButton button * {
            color: #FFFFFF !important;
            -webkit-text-fill-color: #FFFFFF !important;
        }

        section[data-testid="stSidebar"] div[data-baseweb="select"] svg,
        section[data-testid="stSidebar"] div[data-baseweb="input"] svg {
            color: #334155 !important;
            fill: #334155 !important;
        }

        section[data-testid="stSidebar"] [data-baseweb="tag"] {
            background: var(--atlas-chip) !important;
            border: 1px solid rgba(0, 107, 104, 0.22) !important;
            border-radius: 999px !important;
            box-shadow: none !important;
        }

        section[data-testid="stSidebar"] [data-baseweb="tag"] *,
        section[data-testid="stSidebar"] [data-baseweb="tag"] span,
        section[data-testid="stSidebar"] [data-baseweb="tag"] svg {
            color: #12313E !important;
            -webkit-text-fill-color: #12313E !important;
            fill: #12313E !important;
        }

        section[data-testid="stSidebar"] [data-testid="stExpander"] {
            border-radius: 8px !important;
            border: 1px solid rgba(255, 255, 255, 0.72) !important;
            background: rgba(255, 255, 255, 0.07) !important;
            box-shadow: none !important;
        }

        section[data-testid="stSidebar"] [data-testid="stExpander"] summary,
        section[data-testid="stSidebar"] [data-testid="stExpander"] summary * {
            color: #FFFFFF !important;
            -webkit-text-fill-color: #FFFFFF !important;
            font-weight: 760 !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipHoverTarget"],
        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"],
        section[data-testid="stSidebar"] button[aria-label="Help"],
        section[data-testid="stSidebar"] button[title="Help"] {
            position: relative !important;
            width: 18px !important;
            height: 18px !important;
            min-width: 18px !important;
            min-height: 18px !important;
            border-radius: 999px !important;
            background: #063F3C !important;
            border: 1.5px solid #BDF7EF !important;
            box-shadow: 0 0 0 2px rgba(255, 255, 255, 0.08) !important;
            opacity: 1 !important;
            display: inline-flex !important;
            align-items: center !important;
            justify-content: center !important;
            overflow: hidden !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipHoverTarget"] svg,
        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"] svg,
        section[data-testid="stSidebar"] button[aria-label="Help"] svg,
        section[data-testid="stSidebar"] button[title="Help"] svg {
            display: none !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipHoverTarget"]::after,
        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"]::after,
        section[data-testid="stSidebar"] button[aria-label="Help"]::after,
        section[data-testid="stSidebar"] button[title="Help"]::after {
            content: "?" !important;
            color: #FFFFFF !important;
            -webkit-text-fill-color: #FFFFFF !important;
            font: 800 12px/1 Arial, sans-serif !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipHoverTarget"]:hover,
        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"]:hover,
        section[data-testid="stSidebar"] button[aria-label="Help"]:hover,
        section[data-testid="stSidebar"] button[title="Help"]:hover {
            background: var(--atlas-teal-bright) !important;
            border-color: #FFFFFF !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"] [data-testid="stTooltipHoverTarget"] {
            background: transparent !important;
            border: 0 !important;
            box-shadow: none !important;
            overflow: visible !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"] [data-testid="stTooltipHoverTarget"]::after,
        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"] button[aria-label^="Help"]::after {
            content: none !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"] button[aria-label^="Help"] {
            background: transparent !important;
            border: 0 !important;
            box-shadow: none !important;
            padding: 0 !important;
            width: 18px !important;
            height: 18px !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"] button[aria-label^="Help"] svg {
            display: block !important;
            width: 13px !important;
            height: 13px !important;
            color: #FFFFFF !important;
            stroke: #FFFFFF !important;
            fill: none !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"]::after {
            content: none !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"] [data-testid="stTooltipHoverTarget"] {
            position: absolute !important;
            inset: 0 !important;
            width: 100% !important;
            height: 100% !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
        }

        section[data-testid="stSidebar"] [data-testid="stTooltipIcon"] button[aria-label^="Help"] {
            display: inline-flex !important;
            align-items: center !important;
            justify-content: center !important;
            width: 100% !important;
            height: 100% !important;
        }

        .dashboard-hero {
            box-sizing: border-box !important;
            min-height: 164px !important;
            margin: 0 0 0.82rem 0 !important;
            padding: 1.45rem 1.65rem 0.1rem 1.65rem !important;
            border-radius: 13px !important;
            border: 1px solid rgba(15, 23, 42, 0.11) !important;
            background: rgba(255, 255, 255, 0.93) !important;
            box-shadow: 0 2px 12px rgba(15, 23, 42, 0.12) !important;
            backdrop-filter: blur(4px) !important;
        }

        .eyebrow {
            color: var(--atlas-teal) !important;
            -webkit-text-fill-color: var(--atlas-teal) !important;
            font-size: 0.75rem !important;
            font-weight: 750 !important;
            letter-spacing: 0.14em !important;
            text-transform: uppercase !important;
            margin-bottom: 0.82rem !important;
        }

        .dashboard-title {
            color: var(--atlas-ink) !important;
            -webkit-text-fill-color: var(--atlas-ink) !important;
            font-family: "Segoe UI", "Inter", "Aptos Display", sans-serif !important;
            font-size: clamp(2.9rem, 3.6vw, 3.7rem) !important;
            font-weight: 400 !important;
            letter-spacing: 0 !important;
            line-height: 1.04 !important;
            margin: 0 !important;
            padding: 0 !important;
        }

        .dashboard-subtitle {
            color: var(--atlas-muted) !important;
            -webkit-text-fill-color: var(--atlas-muted) !important;
            font-size: 0.96rem !important;
            font-weight: 500 !important;
            margin-top: 0.75rem !important;
        }

        .api-load-caption,
        .atlas-pill {
            display: inline-flex !important;
            align-items: center !important;
            gap: 0.52rem !important;
            min-height: 37px !important;
            margin: 0 0 1.15rem 0 !important;
            padding: 0.4rem 0.9rem !important;
            border-radius: 999px !important;
            border: 1px solid rgba(15, 23, 42, 0.12) !important;
            background: rgba(255, 255, 255, 0.96) !important;
            box-shadow: 0 1px 8px rgba(15, 23, 42, 0.06) !important;
            color: var(--atlas-muted) !important;
            -webkit-text-fill-color: var(--atlas-muted) !important;
            font-size: 0.82rem !important;
            font-weight: 700 !important;
        }

        .api-load-caption strong,
        .api-load-caption span,
        .atlas-pill span {
            color: var(--atlas-ink) !important;
            -webkit-text-fill-color: var(--atlas-ink) !important;
            font-weight: 800 !important;
        }

        .api-load-clock {
            width: 17px;
            height: 17px;
            color: var(--atlas-teal);
            -webkit-text-fill-color: var(--atlas-teal);
            display: inline-flex;
            align-items: center;
            justify-content: center;
        }

        .api-load-clock svg {
            width: 17px;
            height: 17px;
            stroke: currentColor;
            stroke-width: 2.1;
            fill: none;
            stroke-linecap: round;
            stroke-linejoin: round;
        }

        div[data-testid="stTabs"] [data-baseweb="tab-list"] {
            gap: 1.35rem !important;
            border-bottom: 1px solid var(--atlas-line) !important;
        }

        button[data-baseweb="tab"] {
            height: 42px !important;
            padding: 0 0.45rem !important;
            color: #334155 !important;
            -webkit-text-fill-color: #334155 !important;
            font-size: 0.94rem !important;
            font-weight: 500 !important;
            letter-spacing: 0 !important;
        }

        button[data-baseweb="tab"][aria-selected="true"] {
            color: var(--atlas-teal) !important;
            -webkit-text-fill-color: var(--atlas-teal) !important;
            font-weight: 700 !important;
        }

        div[data-baseweb="tab-highlight"] {
            background-color: var(--atlas-teal) !important;
            height: 2px !important;
        }

        .section-title {
            color: var(--atlas-ink) !important;
            -webkit-text-fill-color: var(--atlas-ink) !important;
            font-size: 1.58rem !important;
            line-height: 1.22 !important;
            font-weight: 400 !important;
            letter-spacing: 0 !important;
            margin: 0.55rem 0 0.45rem 0 !important;
        }

        .stApp [data-testid="stMarkdownContainer"] p,
        .stApp [data-testid="stCaptionContainer"] p {
            color: var(--atlas-muted) !important;
            -webkit-text-fill-color: var(--atlas-muted) !important;
        }

        .stApp section[data-testid="stSidebar"] [data-testid="stWidgetLabel"] *,
        .stApp section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p,
        .stApp section[data-testid="stSidebar"] [data-testid="stCaptionContainer"] p {
            color: #FFFFFF !important;
            -webkit-text-fill-color: #FFFFFF !important;
        }

        [data-testid="stRadio"] > label,
        [data-testid="stRadio"] > label * {
            color: var(--atlas-ink) !important;
            -webkit-text-fill-color: var(--atlas-ink) !important;
            font-weight: 650 !important;
        }

        [data-testid="stRadio"] [role="radiogroup"] { gap: 1.9rem !important; }

        [data-testid="stRadio"] [role="radiogroup"] label {
            background: transparent !important;
            border: 0 !important;
            box-shadow: none !important;
            padding: 0 !important;
        }

        [data-testid="stRadio"] [role="radiogroup"] label *,
        [data-testid="stRadio"] [role="radiogroup"] p {
            color: var(--atlas-ink) !important;
            -webkit-text-fill-color: var(--atlas-ink) !important;
            font-weight: 500 !important;
        }

        .atlas-metric-grid {
            display: grid;
            grid-template-columns: repeat(4, minmax(0, 1fr));
            gap: 1.35rem;
            margin: 0 0 0.4rem 0;
        }

        .atlas-metric-card {
            display: grid;
            grid-template-columns: 58px minmax(0, 1fr);
            align-items: center;
            gap: 1rem;
            min-height: 95px;
            padding: 1rem 1.05rem;
            background: #FFFFFF;
            border: 1px solid rgba(15, 23, 42, 0.11);
            border-radius: 9px;
            box-shadow: 0 3px 13px rgba(15, 23, 42, 0.08);
        }

        .atlas-metric-icon {
            width: 55px;
            height: 55px;
            border-radius: 8px;
            background: linear-gradient(135deg, #006F6A 0%, var(--atlas-teal-bright) 100%);
            color: #FFFFFF;
            display: inline-flex;
            align-items: center;
            justify-content: center;
        }

        .atlas-metric-icon svg {
            width: 32px;
            height: 32px;
            stroke: #FFFFFF;
            stroke-width: 2.2;
            fill: none;
            stroke-linecap: round;
            stroke-linejoin: round;
        }

        .atlas-metric-label {
            color: var(--atlas-muted);
            font-size: 0.88rem;
            line-height: 1.2;
            font-weight: 600;
            margin-bottom: 0.35rem;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }

        .atlas-metric-value {
            color: var(--atlas-ink);
            font-size: 1.85rem;
            line-height: 1;
            font-weight: 400;
            letter-spacing: 0;
        }

        .atlas-table-frame,
        div[data-testid="stDataFrame"] {
            border: 1px solid rgba(15, 23, 42, 0.10) !important;
            border-radius: 7px !important;
            background: #FFFFFF !important;
            overflow: hidden;
            box-shadow: 0 2px 10px rgba(15, 23, 42, 0.08) !important;
        }

        .atlas-table-frame { margin-top: 0.4rem; }
        .atlas-table-scroll { width: 100%; overflow: auto; max-height: 560px; }

        .atlas-preview-table {
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
            font-size: 0.88rem;
            color: var(--atlas-ink);
        }

        .atlas-preview-table thead th {
            position: sticky;
            top: 0;
            z-index: 1;
            text-align: left;
            padding: 0.62rem 1rem;
            color: #FFFFFF;
            background: linear-gradient(180deg, #006F6A 0%, #005D59 100%);
            border-right: 1px solid rgba(255, 255, 255, 0.20);
            font-weight: 750;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }

        .atlas-preview-table tbody td {
            padding: 0.48rem 1rem;
            border-top: 1px solid #E7EDEF;
            border-right: 1px solid #E7EDEF;
            background: #FFFFFF;
            color: #12233D;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }

        .atlas-preview-table tbody tr:nth-child(even) td { background: #FCFDFD; }
        .atlas-preview-table th:last-child,
        .atlas-preview-table td:last-child { border-right: 0; }

        .atlas-table-empty {
            margin-top: 1rem;
            padding: 1.1rem 1.2rem;
            border: 1px solid rgba(15, 23, 42, 0.10);
            border-radius: 9px;
            background: #FFFFFF;
            color: var(--atlas-muted);
            box-shadow: 0 2px 10px rgba(15, 23, 42, 0.06);
        }

        div[data-testid="stMetric"] {
            background: #FFFFFF !important;
            border: 1px solid rgba(15, 23, 42, 0.10) !important;
            border-radius: 9px !important;
            box-shadow: 0 2px 10px rgba(15, 23, 42, 0.08) !important;
        }

        [data-baseweb="popover"],
        [data-baseweb="menu"],
        [role="listbox"] {
            background: #FFFFFF !important;
            color: var(--atlas-ink) !important;
        }

        [role="option"],
        [role="option"] *,
        [data-baseweb="menu"] * {
            color: var(--atlas-ink) !important;
            -webkit-text-fill-color: var(--atlas-ink) !important;
        }

        div[data-testid="stAlert"],
        div[data-testid="stAlert"] > div,
        div[data-testid="stAlert"] [role="alert"] {
            background: rgba(255, 255, 255, 0.92) !important;
            border: 1px solid rgba(15, 118, 110, 0.18) !important;
            color: var(--atlas-ink) !important;
            border-radius: 9px !important;
            box-shadow: none !important;
        }

        .stButton button,
        .stDownloadButton button,
        div[data-testid="stButton"] button,
        div[data-testid="stDownloadButton"] button,
        button[kind="primary"],
        button[kind="secondary"] {
            border-radius: 8px !important;
            background: linear-gradient(135deg, #006F6A, var(--atlas-teal-bright)) !important;
            border: 1px solid rgba(0, 107, 104, 0.32) !important;
            box-shadow: none !important;
        }

        .stButton button,
        .stButton button *,
        .stDownloadButton button,
        .stDownloadButton button *,
        div[data-testid="stButton"] button,
        div[data-testid="stButton"] button *,
        div[data-testid="stDownloadButton"] button,
        div[data-testid="stDownloadButton"] button *,
        button[kind="primary"],
        button[kind="primary"] *,
        button[kind="secondary"],
        button[kind="secondary"] * {
            color: #FFFFFF !important;
            -webkit-text-fill-color: #FFFFFF !important;
            font-weight: 600 !important;
            opacity: 1 !important;
            text-shadow: none !important;
        }

        .stButton button:hover,
        .stDownloadButton button:hover,
        div[data-testid="stButton"] button:hover,
        div[data-testid="stDownloadButton"] button:hover,
        button[kind="primary"]:hover,
        button[kind="secondary"]:hover {
            color: #FFFFFF !important;
            -webkit-text-fill-color: #FFFFFF !important;
            filter: brightness(1.04);
        }

        @media (max-width: 1100px) {
            .atlas-metric-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
        }

        @media (max-width: 640px) {
            .atlas-metric-grid { grid-template-columns: 1fr; }
            .dashboard-title { font-size: 2.55rem !important; }
            .dashboard-hero { min-height: 150px !important; }
        }
        </style>
        <div class="atlas-topbar-brand" aria-hidden="true">
            <div class="atlas-menu-lines"><span></span></div>
            <div class="atlas-logo-mark"></div>
            <div class="atlas-brand-word">AtlasFlow</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_header(selected_group: str, selected_vessels: list[str], selected_variables: list[str]) -> None:
    vessel_text = "All selected vessels" if len(selected_vessels) != 1 else selected_vessels[0]
    variable_text = "No variables selected" if not selected_variables else f"{len(selected_variables):,} selected variables"
    st.markdown(
        f"""
        <div class="dashboard-hero">
            <div class="eyebrow">Marorka API Explorer</div>
            <h1 class="dashboard-title">Atlas Flow</h1>
            <div class="dashboard-subtitle">
                {escape(selected_group)} | {escape(vessel_text)} | {escape(variable_text)}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_api_load_caption(metadata: dict[str, Any] | None) -> None:
    metadata = metadata or {}
    last_load = metadata.get("loaded_at_local") or metadata.get("loaded_at_utc") or "-"
    last_load_display = str(last_load).replace(" EEST", "").replace(" EET", "")
    st.markdown(
        f"""
        <div class="api-load-caption">
            <span class="api-load-clock" aria-hidden="true">
                <svg viewBox="0 0 24 24">
                    <circle cx="12" cy="12" r="9"></circle>
                    <path d="M12 7v5l3 2"></path>
                </svg>
            </span>
            <strong>Last API load:</strong> <span>{escape(last_load_display)} LT</span>
        </div>
        """,
        unsafe_allow_html=True,
    )


# =============================================================================
# Secrets/auth/API helpers
# =============================================================================


class MarorkaConfigError(RuntimeError):
    pass


def read_secret(name: str, default: str = "") -> str:
    try:
        value = st.secrets.get(name, os.getenv(name, default))
    except Exception:
        value = os.getenv(name, default)
    return str(value).strip() if value is not None else default


def app_timezone() -> ZoneInfo:
    timezone_name = read_secret("APP_TIMEZONE", "Europe/Athens")
    try:
        return ZoneInfo(timezone_name)
    except Exception:
        return ZoneInfo("Europe/Athens")


def local_time_label(dt_utc: datetime | None = None) -> str:
    dt_utc = dt_utc or datetime.now(timezone.utc)
    if dt_utc.tzinfo is None:
        dt_utc = dt_utc.replace(tzinfo=timezone.utc)
    local_dt = dt_utc.astimezone(app_timezone())
    return local_dt.strftime("%d-%m-%Y %H:%M:%S %Z")


def get_query_param(name: str, default: str = "") -> str:
    try:
        value = st.query_params.get(name, default)
    except Exception:
        value = st.experimental_get_query_params().get(name, [default])
    if isinstance(value, list):
        return str(value[0]) if value else default
    return str(value) if value is not None else default


def is_warmup_request() -> bool:
    return get_query_param("warmup", "0") == "1"


def warmup_token_is_valid() -> bool:
    expected_token = read_secret("WARMUP_TOKEN")
    provided_token = get_query_param("token", "")
    return bool(expected_token) and hmac.compare_digest(provided_token, expected_token)


def require_dashboard_password() -> None:
    dashboard_password = read_secret("DASHBOARD_PASSWORD")
    if not dashboard_password:
        return

    if st.session_state.get("dashboard_authenticated"):
        return

    apply_custom_css()
    st.markdown(
        """
        <div class="dashboard-hero">
            <div class="eyebrow">Secure access</div>
            <h1 class="dashboard-title">Atlas Flow</h1>
            <div class="dashboard-subtitle">Enter your dashboard password to continue.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    entered_password = st.text_input("Password", type="password")

    if st.button("Sign in", type="primary"):
        if hmac.compare_digest(entered_password, dashboard_password):
            st.session_state["dashboard_authenticated"] = True
            st.rerun()
        st.error("Invalid password.")

    st.stop()


def request_auth(username: str, password: str, auth_method: str) -> Any:
    method = auth_method.lower()
    if method == "basic":
        return HTTPBasicAuth(username, password)
    if method == "digest":
        return HTTPDigestAuth(username, password)
    if method == "bearer":
        return None
    if method in {"none", "anonymous", ""}:
        return None
    raise MarorkaConfigError("Unsupported MARORKA_AUTH_METHOD. Use basic, digest, bearer, or none.")


def request_headers(token: str, auth_method: str) -> dict[str, str]:
    headers = {"Accept": "application/json"}
    if auth_method.lower() == "bearer":
        if not token:
            raise MarorkaConfigError("MARORKA_TOKEN is required for bearer auth.")
        headers["Authorization"] = f"Bearer {token}"
    return headers


def odata_quote(value: str) -> str:
    return str(value).replace("'", "''")


def build_reportdata_value_filter() -> str:
    value_filters = [
        f"ValueDescription eq '{odata_quote(value)}'"
        for value in REPORTDATA_VALUE_WHITELIST
    ]
    return "(" + " or ".join(value_filters) + ")"


def build_odata_url(start_date: date) -> str:
    # Keep the OData request simple. The Marorka OData V1 ReportData endpoint
    # rejects long ValueDescription OR filters with 404. We therefore request
    # the date window only, then apply the KPI/consumption ValueDescription
    # whitelist locally inside compact_odata_rows() page-by-page before writing
    # to the Parquet snapshot. This preserves the same final dataset while
    # avoiding invalid/oversized OData URLs.
    start_text = start_date.strftime("%Y-%m-%d")
    params = {
        "$filter": f"StartDateTimeGMT gt DateTime'{start_text}'",
        "$select": ",".join(SOURCE_COLUMNS),
    }
    return f"{ODATA_ENDPOINT}?{urlencode(params)}"


def extract_odata_page(payload: Any) -> tuple[list[dict[str, Any]], str | None]:
    if isinstance(payload, list):
        return payload, None

    if not isinstance(payload, dict):
        raise ValueError("Could not parse OData response payload.")

    rows = payload.get("value")
    next_link = payload.get("@odata.nextLink") or payload.get("odata.nextLink")

    if rows is None and isinstance(payload.get("d"), dict):
        data = payload["d"]
        rows = data.get("results")
        next_link = next_link or data.get("__next")

    if rows is None:
        raise ValueError("Could not find OData rows in the API response.")

    return rows, next_link


def rows_to_dataframe(rows: list[dict[str, Any]]) -> pd.DataFrame:
    df = pd.DataFrame(rows)
    if "__metadata" in df.columns:
        df = df.drop(columns=["__metadata"])
    for column in SOURCE_COLUMNS:
        if column not in df.columns:
            df[column] = pd.NA
    return df[SOURCE_COLUMNS]


def compact_odata_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    compact_rows: list[dict[str, Any]] = []
    for row in rows:
        value_description = row.get("ValueDescription")
        if value_description is None:
            continue
        if re.sub(r"[^a-z0-9]+", "", str(value_description).lower()) not in REPORTDATA_VALUE_WHITELIST_KEYS:
            continue
        if row.get("ReportType") in EXCLUDED_REPORT_TYPES:
            continue
        compact_rows.append({column: row.get(column) for column in SOURCE_COLUMNS})
    return compact_rows


def fetch_report_data(
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    started_at = time.perf_counter()
    next_url = build_odata_url(start_date)
    kept_rows: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    pages = 0
    total_bytes = 0
    scanned_rows = 0
    first_url = next_url
    auth = request_auth(username, password, auth_method)
    headers = request_headers(token, auth_method)

    with requests.Session() as session:
        session.headers.update(headers)
        for _ in range(MAX_ODATA_PAGES):
            if next_url in seen_urls:
                break
            seen_urls.add(next_url)

            response = session.get(next_url, auth=auth, timeout=90)
            total_bytes += len(response.content)
            response.raise_for_status()
            pages += 1

            page_rows, next_link = extract_odata_page(response.json())
            scanned_rows += len(page_rows)
            kept_rows.extend(compact_odata_rows(page_rows))

            if not next_link:
                break
            next_url = urljoin(next_url, next_link)

    loaded_at_utc = datetime.now(timezone.utc)
    metadata = {
        "loaded_at_utc": loaded_at_utc.strftime("%d-%m-%Y %H:%M:%S UTC"),
        "loaded_at_local": local_time_label(loaded_at_utc),
        "rows": len(kept_rows),
        "kept_rows": len(kept_rows),
        "scanned_rows": scanned_rows,
        "discarded_rows": max(scanned_rows - len(kept_rows), 0),
        "pages": pages,
        "downloaded_mb": round(total_bytes / 1024 / 1024, 2),
        "fetch_seconds": round(time.perf_counter() - started_at, 2),
        "first_url": first_url,
        "hit_page_limit": pages >= MAX_ODATA_PAGES,
    }
    return rows_to_dataframe(kept_rows), metadata


@st.cache_data(ttl=API_CACHE_TTL_SECONDS, show_spinner=False)
def cached_fetch_report_data(
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    return fetch_report_data(username, password, token, auth_method, start_date)


# =============================================================================
# Multi-source wide OData helpers
# =============================================================================


def build_wide_odata_url(endpoint: str, start_date: date, datetime_column: str = "DateTime") -> str:
    start_text = start_date.strftime("%Y-%m-%d")
    params = {"$filter": f"{datetime_column} gt DateTime'{start_text}'"}
    return f"{endpoint}?{urlencode(params)}"


def fetch_wide_odata_source(
    source_key: str,
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    config = SOURCE_CONFIGS[source_key]
    endpoint = str(config["endpoint"])
    datetime_column = str(config.get("datetime_candidates", ["DateTime"])[0])
    next_url = build_wide_odata_url(endpoint, start_date, datetime_column)
    first_url = next_url
    seen_urls: set[str] = set()
    rows: list[dict[str, Any]] = []
    pages = 0
    total_bytes = 0
    started_at = time.perf_counter()
    auth = request_auth(username, password, auth_method)
    headers = request_headers(token, auth_method)

    with requests.Session() as session:
        session.headers.update(headers)
        for _ in range(MAX_ODATA_PAGES):
            if next_url in seen_urls:
                break
            seen_urls.add(next_url)
            response = session.get(next_url, auth=auth, timeout=90)
            total_bytes += len(response.content)
            response.raise_for_status()
            pages += 1
            page_rows, next_link = extract_odata_page(response.json())
            rows.extend(page_rows)
            if not next_link:
                break
            next_url = urljoin(next_url, next_link)

    df = pd.DataFrame(rows)
    if "__metadata" in df.columns:
        df = df.drop(columns=["__metadata"])

    loaded_at_utc = datetime.now(timezone.utc)
    metadata = {
        "source": config["label"],
        "endpoint": endpoint,
        "loaded_at_utc": loaded_at_utc.strftime("%d-%m-%Y %H:%M:%S UTC"),
        "loaded_at_local": local_time_label(loaded_at_utc),
        "loaded_start_date": start_date.isoformat(),
        "rows": int(len(df)),
        "columns": int(len(df.columns)),
        "pages": pages,
        "downloaded_mb": round(total_bytes / 1024 / 1024, 2),
        "fetch_seconds": round(time.perf_counter() - started_at, 2),
        "first_url": first_url,
        "hit_page_limit": pages >= MAX_ODATA_PAGES,
    }
    return df, metadata


@st.cache_data(ttl=API_CACHE_TTL_SECONDS, show_spinner=False)
def cached_fetch_wide_odata_source(
    source_key: str,
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    return fetch_wide_odata_source(source_key, username, password, token, auth_method, start_date)


def source_signature(source_key: str, username: str, auth_method: str, start_date: date) -> dict[str, Any]:
    config = SOURCE_CONFIGS[source_key]
    return {
        "source": source_key,
        "endpoint": str(config["endpoint"]),
        "username_hash": sha256(username.encode("utf-8")).hexdigest()[:12],
        "auth_method": auth_method.lower(),
        "start_date": start_date.isoformat(),
    }


def save_source_snapshot(source_key: str, df: pd.DataFrame, metadata: dict[str, Any], signature: dict[str, Any]) -> None:
    try:
        config = SOURCE_CONFIGS[source_key]
        SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
        df.to_parquet(config["snapshot_file"], index=False)
        payload = {
            "metadata": metadata,
            "signature": signature,
            "saved_at_utc": datetime.now(timezone.utc).strftime("%d-%m-%Y %H:%M:%S UTC"),
        }
        Path(config["metadata_file"]).write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    except Exception:
        return


def load_source_snapshot(
    source_key: str,
    requested_signature: dict[str, Any],
    requested_start_date: date,
) -> tuple[pd.DataFrame, dict[str, Any], dict[str, Any]] | None:
    try:
        config = SOURCE_CONFIGS[source_key]
        snapshot_file = Path(config["snapshot_file"])
        metadata_file = Path(config["metadata_file"])
        if not snapshot_file.is_file() or not metadata_file.is_file():
            return None
        payload = json.loads(metadata_file.read_text(encoding="utf-8"))
        metadata = payload.get("metadata") or {}
        signature = payload.get("signature") or {}
        if not raw_data_covers_request(signature, metadata, requested_signature, requested_start_date):
            return None
        df = pd.read_parquet(snapshot_file)
        if not isinstance(df, pd.DataFrame):
            return None
        # Reject broken placeholder snapshots that can be created after an interrupted warmup.
        # If API metadata says rows exist, a one-column NoData parquet is not a usable source snapshot.
        if list(df.columns) == ["NoData"] and int(metadata.get("rows", 0) or 0) > 0:
            return None
        metadata = metadata.copy()
        metadata["loaded_from_snapshot"] = True
        metadata.setdefault("snapshot_saved_at_utc", payload.get("saved_at_utc", "-"))
        return df, metadata, signature
    except Exception:
        return None


def parse_wide_source_datetimes(df: pd.DataFrame) -> pd.DataFrame:
    parsed_df = df.copy()
    for column in parsed_df.columns:
        if "date" in str(column).lower() or "time" in str(column).lower():
            parsed = parse_datetime_series(parsed_df[column])
            if parsed.notna().any():
                parsed_df[column] = parsed
    return parsed_df


def detect_datetime_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for column in candidates:
        if column in df.columns:
            return column
    for column in df.columns:
        lower = str(column).lower()
        if "datetime" in lower or lower in {"date", "timestamp"}:
            return column
    return None


def filter_wide_source_data(
    df: pd.DataFrame,
    source_key: str,
    selected_vessels: list[str],
    selected_start: date,
    selected_end: date,
) -> pd.DataFrame:
    if df.empty:
        return df
    filtered = parse_wide_source_datetimes(df)
    if "ShipName" in filtered.columns and selected_vessels:
        filtered = filtered[match_selected_vessels(filtered["ShipName"], selected_vessels)].copy()
    datetime_column = detect_datetime_column(filtered, list(SOURCE_CONFIGS[source_key].get("datetime_candidates", [])))
    if datetime_column and datetime_column in filtered.columns:
        values = pd.to_datetime(filtered[datetime_column], errors="coerce", utc=True)
        start_timestamp = pd.Timestamp(selected_start, tz="UTC")
        end_timestamp = pd.Timestamp(selected_end + timedelta(days=1), tz="UTC")
        filtered = filtered[values.ge(start_timestamp) & values.lt(end_timestamp)].copy()
    return filtered


def load_or_fetch_source(
    source_key: str,
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
    refresh: bool,
    auto_fetch: bool = False,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    sig = source_signature(source_key, username, auth_method, start_date)
    state_df_key = f"loaded_{source_key}_df"
    state_meta_key = f"loaded_{source_key}_metadata"
    state_sig_key = f"loaded_{source_key}_signature"
    df = st.session_state.get(state_df_key)
    metadata = st.session_state.get(state_meta_key)
    current_signature = st.session_state.get(state_sig_key)

    needs_load = (
        refresh
        or not isinstance(df, pd.DataFrame)
        or not isinstance(metadata, dict)
        or not raw_data_covers_request(current_signature, metadata, sig, start_date)
    )

    if needs_load and not refresh:
        snapshot = load_source_snapshot(source_key, sig, start_date)
        if snapshot is not None:
            df, metadata, snapshot_sig = snapshot
            st.session_state[state_df_key] = df
            st.session_state[state_meta_key] = metadata
            st.session_state[state_sig_key] = snapshot_sig
            needs_load = False

    if needs_load and not (refresh or auto_fetch):
        config = SOURCE_CONFIGS[source_key]
        empty_metadata = {
            "source": config["label"],
            "endpoint": str(config["endpoint"]),
            "loaded_at_utc": "-",
            "loaded_at_local": "No stored snapshot yet",
            "loaded_from_snapshot": False,
            "rows": 0,
            "columns": 0,
            "pages": 0,
            "first_url": "-",
            "needs_warmup": True,
        }
        return pd.DataFrame(), empty_metadata

    if needs_load:
        if refresh:
            cached_fetch_wide_odata_source.clear()
        df, metadata = fetch_wide_odata_source(source_key, username, password, token, auth_method, start_date)
        save_source_snapshot(source_key, df, metadata, sig)
        st.session_state[state_df_key] = df
        st.session_state[state_meta_key] = metadata
        st.session_state[state_sig_key] = sig

    return st.session_state[state_df_key], st.session_state[state_meta_key]


def render_wide_source_tab(source_label: str, df: pd.DataFrame, metadata: dict[str, Any], source_key: str, selected_vessels: list[str], selected_start: date, selected_end: date) -> pd.DataFrame:
    st.markdown(f'<div class="section-title">{escape(source_label)} Dataset</div>', unsafe_allow_html=True)
    render_api_load_caption(metadata)
    if list(df.columns) == ["NoData"] and int(metadata.get("rows", 0) or 0) > 0:
        st.error(
            f"{source_label} snapshot metadata shows {int(metadata.get('rows', 0)):,} API rows, "
            "but the stored parquet contains only a placeholder column. "
            f"Run the {source_key} warmup again with the latest app version."
        )
        return pd.DataFrame()

    filtered_df = filter_wide_source_data(df, source_key, selected_vessels, selected_start, selected_end)
    render_metric_cards(
        [
            ("Rows", f"{len(filtered_df):,}", "table_eye"),
            ("Columns", f"{len(filtered_df.columns):,}", "checked_columns"),
            ("API Rows", f"{metadata.get('rows', len(df)):,}", "database_rows"),
            ("API Pages", f"{metadata.get('pages', 0):,}", "numeric"),
        ]
    )

    default_columns = [c for c in ["ShipName", "DateTime", "State", "StateName", "GPSSpeed", "LogSpeed", "MEConsumed", "ShaftPower"] if c in filtered_df.columns]
    if not default_columns:
        default_columns = list(filtered_df.columns[: min(12, len(filtered_df.columns))])
    selected_columns = st.multiselect(
        f"{source_label} columns to preview/export",
        options=list(filtered_df.columns),
        default=default_columns,
        key=f"{source_key}_preview_columns",
    )
    if not selected_columns:
        selected_columns = default_columns
    output = filtered_df[selected_columns].copy() if selected_columns else filtered_df.copy()
    render_preview_table(output)
    if len(output) > TABLE_PREVIEW_ROW_LIMIT:
        st.caption(f"Showing first {TABLE_PREVIEW_ROW_LIMIT:,} of {len(output):,} rows. Export includes all filtered rows/columns selected above.")
    return output


def to_multisource_excel_bytes(
    reportdata_df: pd.DataFrame,
    reportdata_summary_df: pd.DataFrame | None,
    reportpivots_df: pd.DataFrame,
    shippivots_df: pd.DataFrame,
) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        write_table_sheet(writer, reportdata_df, "ReportData Clean", "ReportDataClean")
        if reportdata_summary_df is not None and not reportdata_summary_df.empty:
            write_table_sheet(writer, reportdata_summary_df, "ReportData Summary", "ReportDataSummary")
        if reportpivots_df is not None and not reportpivots_df.empty:
            write_table_sheet(writer, reportpivots_df, "ReportPivots", "ReportPivotsData")
        if shippivots_df is not None and not shippivots_df.empty:
            write_table_sheet(writer, shippivots_df, "ShipPivots", "ShipPivotsData")
    return output.getvalue()


# =============================================================================
# Transform helpers
# =============================================================================


def normalize_text(value: Any) -> str:
    text = str(value).lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def parse_datetime_series(series: pd.Series) -> pd.Series:
    parsed = pd.to_datetime(series, errors="coerce", utc=True)
    missing_mask = parsed.isna()

    if missing_mask.any():
        date_text = series.astype("string")
        dotnet_millis = date_text.str.extract(r"/Date\((-?\d+)").iloc[:, 0]
        dotnet_parsed = pd.to_datetime(
            pd.to_numeric(dotnet_millis, errors="coerce"),
            errors="coerce",
            unit="ms",
            utc=True,
        )
        parsed = parsed.mask(missing_mask, dotnet_parsed)

    return parsed


def parse_numeric_value(value: Any) -> Any:
    if pd.isna(value):
        return pd.NA
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return pd.NA

    duration_match = re.fullmatch(r"(-?\d+):([0-5]?\d)(?::([0-5]?\d))?", text)
    if duration_match:
        hours = int(duration_match.group(1))
        sign = -1 if hours < 0 else 1
        minutes = int(duration_match.group(2))
        seconds = int(duration_match.group(3) or 0)
        return sign * (abs(hours) + minutes / 60 + seconds / 3600)

    numeric_text = text.replace(" ", "")
    if re.fullmatch(r"-?\d+,\d+", numeric_text):
        numeric_text = numeric_text.replace(",", ".")
    else:
        numeric_text = numeric_text.replace(",", "")

    numeric_text = re.sub(r"[^0-9.\-]", "", numeric_text)
    if numeric_text in {"", "-", ".", "-."}:
        return pd.NA

    try:
        return float(numeric_text)
    except ValueError:
        return pd.NA


def parse_numeric_series(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series.map(parse_numeric_value), errors="coerce")


def match_selected_vessels(raw_ship_names: pd.Series, selected_vessels: list[str]) -> pd.Series:
    selected_keys = {normalize_text(vessel) for vessel in selected_vessels}
    return raw_ship_names.map(normalize_text).isin(selected_keys)


@st.cache_data(ttl=API_CACHE_TTL_SECONDS, show_spinner=False)
def cached_prepare_long_data(raw_df: pd.DataFrame) -> pd.DataFrame:
    missing_columns = sorted(set(SOURCE_COLUMNS).difference(raw_df.columns))
    if missing_columns:
        raise ValueError(f"Missing expected API columns: {', '.join(missing_columns)}")

    df = raw_df.copy()
    df["StartDateTimeGMT"] = parse_datetime_series(df["StartDateTimeGMT"])
    df["EndDateTimeGMT"] = parse_datetime_series(df["EndDateTimeGMT"])
    df["LapTime"] = parse_numeric_series(df["LapTime"])
    df["ParsedValue"] = parse_numeric_series(df["ReportedValue"])
    df = df[df["ValueDescription"].notna() & ~df["ReportType"].isin(EXCLUDED_REPORT_TYPES)].copy()
    return df


def available_variables(df: pd.DataFrame) -> list[str]:
    if df.empty or "ValueDescription" not in df.columns:
        return []
    return sorted(df["ValueDescription"].dropna().astype(str).unique().tolist(), key=str.casefold)


def available_report_types(df: pd.DataFrame) -> list[str]:
    if df.empty or "ReportType" not in df.columns:
        return []
    return sorted(df["ReportType"].dropna().astype(str).unique().tolist(), key=str.casefold)


def dataframe_date_window(df: pd.DataFrame) -> tuple[date, date]:
    if df.empty or "StartDateTimeGMT" not in df.columns:
        today = date.today()
        return today, today
    dates = pd.to_datetime(df["StartDateTimeGMT"], errors="coerce", utc=True).dt.date.dropna()
    if dates.empty:
        today = date.today()
        return today, today
    return max(dates.min(), API_FULL_START_DATE), min(dates.max(), date.today())


def filter_long_data(
    df: pd.DataFrame,
    selected_vessels: list[str],
    selected_report_types: list[str],
    selected_start: date,
    selected_end: date,
) -> pd.DataFrame:
    if df.empty:
        return df

    start_timestamp = pd.Timestamp(selected_start, tz="UTC")
    end_timestamp = pd.Timestamp(selected_end + timedelta(days=1), tz="UTC")
    start_values = pd.to_datetime(df["StartDateTimeGMT"], errors="coerce", utc=True)

    filtered = df[
        match_selected_vessels(df["ShipName"], selected_vessels)
        & start_values.ge(start_timestamp)
        & start_values.lt(end_timestamp)
    ].copy()

    if selected_report_types:
        filtered = filtered[filtered["ReportType"].astype("string").isin(selected_report_types)].copy()

    return filtered


def safe_divide(numerator: Any, denominator: Any) -> Any:
    numerator = pd.to_numeric(numerator, errors="coerce")
    denominator = pd.to_numeric(denominator, errors="coerce")
    denominator = denominator.mask(denominator == 0)
    return numerator / denominator


def sum_numeric_columns(df: pd.DataFrame, columns: list[str]) -> pd.Series:
    available_columns = [column for column in columns if column in df.columns]
    if not available_columns:
        return pd.Series(pd.NA, index=df.index, dtype="Float64")
    return df[available_columns].apply(pd.to_numeric, errors="coerce").sum(axis=1, min_count=1)


def calculation_alias_to_column() -> dict[str, str]:
    return {
        normalize_text(alias): column
        for column, aliases in DERIVED_VALUE_ALIASES.items()
        for alias in aliases
    }


def calculate_rob_consumption(
    df: pd.DataFrame,
    rob_column: str,
    received_column: str,
) -> pd.Series:
    """Calculate row-level consumption from ROB movement inside the current sample.

    The first report per vessel has no previous ROB inside the selected sample, so it
    remains blank. Negative values are treated as blank because they usually indicate
    correction/noise or a missing receipt entry rather than real consumption.
    """
    if rob_column not in df.columns:
        return pd.Series(pd.NA, index=df.index, dtype="Float64")

    work = df[["ShipName", "StartDateTimeGMT", "EndDateTimeGMT", rob_column]].copy()
    if received_column in df.columns:
        work[received_column] = pd.to_numeric(df[received_column], errors="coerce").fillna(0)
    else:
        work[received_column] = 0.0

    work[rob_column] = pd.to_numeric(work[rob_column], errors="coerce")
    work["_original_index"] = df.index
    work["_sort_date"] = pd.to_datetime(work["EndDateTimeGMT"], errors="coerce", utc=True)
    fallback_dates = pd.to_datetime(work["StartDateTimeGMT"], errors="coerce", utc=True)
    work["_sort_date"] = work["_sort_date"].fillna(fallback_dates)
    work = work.sort_values(["ShipName", "_sort_date", "_original_index"])
    previous_rob = work.groupby("ShipName", dropna=False)[rob_column].shift(1)
    consumption = previous_rob + work[received_column] - work[rob_column]
    consumption = consumption.where(consumption >= 0)
    consumption = consumption.where(previous_rob.notna() & work[rob_column].notna())
    result = pd.Series(pd.NA, index=df.index, dtype="Float64")
    result.loc[work["_original_index"]] = pd.to_numeric(consumption, errors="coerce").to_numpy()
    return result


def build_calculation_source_table(filtered_long_df: pd.DataFrame) -> pd.DataFrame:
    if filtered_long_df.empty:
        return pd.DataFrame(columns=PIVOT_IDENTITY_COLUMNS)

    alias_to_column = calculation_alias_to_column()
    source_long = filtered_long_df.copy()
    source_long["_value_key"] = source_long["ValueDescription"].map(normalize_text)
    source_long = source_long[
        source_long["_value_key"].isin(alias_to_column)
        & source_long["ParsedValue"].notna()
    ].copy()

    if source_long.empty:
        return filtered_long_df[PIVOT_IDENTITY_COLUMNS].drop_duplicates().copy()

    source_long["_canonical_column"] = source_long["_value_key"].map(alias_to_column)
    source_long["_source_order"] = range(len(source_long))
    source_long = source_long.sort_values("_source_order")
    source_long = source_long.drop_duplicates(
        [*PIVOT_IDENTITY_COLUMNS, "_canonical_column"],
        keep="last",
    )

    source_table = (
        source_long
        .pivot(index=PIVOT_IDENTITY_COLUMNS, columns="_canonical_column", values="ParsedValue")
        .reset_index()
    )
    source_table.columns.name = None
    return source_table


def add_performance_calculations(pivot_df: pd.DataFrame, source_table: pd.DataFrame) -> pd.DataFrame:
    if pivot_df.empty:
        for column in DERIVED_VARIABLES:
            if column not in pivot_df.columns:
                pivot_df[column] = pd.NA
        return pivot_df

    df = pivot_df.copy()
    if not source_table.empty:
        calculation_columns = [
            column for column in source_table.columns
            if column not in PIVOT_IDENTITY_COLUMNS and column not in df.columns
        ]
        if calculation_columns:
            df = df.merge(
                source_table[[*PIVOT_IDENTITY_COLUMNS, *calculation_columns]],
                on=PIVOT_IDENTITY_COLUMNS,
                how="left",
            )

    def numeric_column(column: str) -> pd.Series:
        if column not in df.columns:
            return pd.Series(pd.NA, index=df.index, dtype="Float64")
        return pd.to_numeric(df[column], errors="coerce")

    lap_time = numeric_column("LapTime")
    engine_distance = numeric_column("Engine Distance [nm]")
    distance_over_ground = numeric_column("Distance Over Ground [nm]")
    power = numeric_column("Power from Torque Meter [kW]")

    df["Calculated Slip"] = (1 - safe_divide(distance_over_ground, engine_distance)).round(3)

    me_sum = sum_numeric_columns(df, ME_FUEL_COLUMNS)
    official_me_total = numeric_column("Main Engine Total Consumed")
    df["ME Consumption Total"] = me_sum.fillna(official_me_total).round(3)

    dg_sum = sum_numeric_columns(df, DG_FUEL_COLUMNS)
    official_dg_total = numeric_column("Diesel Generator Total Consumed")
    df["DG Consumption Total"] = dg_sum.fillna(official_dg_total).round(3)

    aux_sum = sum_numeric_columns(df, AUXILIARY_FUEL_COLUMNS)
    official_aux_total = numeric_column("Auxiliary Engine Total Consumed")
    df["Auxiliary Engine Consumption Total"] = aux_sum.fillna(official_aux_total).round(3)

    df["Boiler Sum"] = sum_numeric_columns(df, BOILER_FUEL_COLUMNS).round(3)

    calculated_total = pd.concat(
        [
            pd.to_numeric(df["ME Consumption Total"], errors="coerce"),
            pd.to_numeric(df["DG Consumption Total"], errors="coerce"),
            pd.to_numeric(df["Auxiliary Engine Consumption Total"], errors="coerce"),
            pd.to_numeric(df["Boiler Sum"], errors="coerce"),
        ],
        axis=1,
    ).sum(axis=1, min_count=1)
    official_total = numeric_column("Total Fuel Consumed")
    df["Total Fuel Consumption"] = calculated_total.fillna(official_total).round(3)

    df["Consumption ME 24 Hours [MT]"] = safe_divide(df["ME Consumption Total"] * 24, lap_time).round(3)

    df["SFOC [gr/Kwh]"] = (
        safe_divide(df["Consumption ME 24 Hours [MT]"], power) / 0.000024
    ).round(3).fillna(0)

    # Oil consumption aggregation variables. These are row-level consumption
    # movements calculated inside the current selected sample, so Summary Analysis
    # can later Sum them by vessel/fleet/month/report type.
    df["MELO Consumption Total [ltr]"] = calculate_rob_consumption(
        df,
        "MELO ROB [ltr]",
        "MELO Received [ltr]",
    ).round(3)
    cylinder_oil_1_consumption = calculate_rob_consumption(
        df,
        "Cylinder Oil 1 ROB [ltr]",
        "Cylinder Oil 1 Received [ltr]",
    )
    cylinder_oil_2_consumption = calculate_rob_consumption(
        df,
        "Cylinder Oil 2 ROB [ltr]",
        "Cylinder Oil 2 Received [ltr]",
    )
    df["CYLO Consumption Total [ltr]"] = pd.concat(
        [cylinder_oil_1_consumption, cylinder_oil_2_consumption],
        axis=1,
    ).sum(axis=1, min_count=1).round(3)
    df["GELO Consumption Total [ltr]"] = calculate_rob_consumption(
        df,
        "GELO ROB [ltr]",
        "GELO Received [ltr]",
    ).round(3)
    df["Total DG Running Hours [hh:mm]"] = sum_numeric_columns(
        df,
        [
            "DG1 Running Hours [hh:mm]",
            "DG2 Running Hours [hh:mm]",
            "DG3 Running Hours [hh:mm]",
            "DG4 Running Hours [hh:mm]",
        ],
    ).round(3)

    return df


@st.cache_data(show_spinner=False)
def build_pivot_table(filtered_long_df: pd.DataFrame, selected_variables: tuple[str, ...]) -> pd.DataFrame:
    if filtered_long_df.empty:
        return pd.DataFrame(columns=PIVOT_IDENTITY_COLUMNS + list(selected_variables))

    calculation_source_table = build_calculation_source_table(filtered_long_df)

    api_selected_variables = [
        variable for variable in selected_variables
        if variable not in DERIVED_VARIABLES
    ]

    if not api_selected_variables:
        pivot_df = (
            filtered_long_df[PIVOT_IDENTITY_COLUMNS]
            .drop_duplicates()
            .sort_values(["ShipName", "EndDateTimeGMT"], ascending=[True, False])
            .reset_index(drop=True)
        )
    else:
        selected_long = filtered_long_df[
            filtered_long_df["ValueDescription"].astype("string").isin(api_selected_variables)
        ].copy()

        if selected_long.empty:
            pivot_df = filtered_long_df[PIVOT_IDENTITY_COLUMNS].drop_duplicates().copy()
            for variable in api_selected_variables:
                pivot_df[variable] = pd.NA
        else:
            selected_long["ValueDescription"] = selected_long["ValueDescription"].astype(str)
            selected_long["_source_order"] = range(len(selected_long))
            selected_long = selected_long.sort_values("_source_order")
            selected_long = selected_long.drop_duplicates(
                [*PIVOT_IDENTITY_COLUMNS, "ValueDescription"],
                keep="last",
            )

            pivot_df = (
                selected_long
                .pivot(index=PIVOT_IDENTITY_COLUMNS, columns="ValueDescription", values="ParsedValue")
                .reset_index()
            )
            pivot_df.columns.name = None

        for variable in api_selected_variables:
            if variable not in pivot_df.columns:
                pivot_df[variable] = pd.NA

    pivot_df = add_performance_calculations(pivot_df, calculation_source_table)

    for variable in selected_variables:
        if variable not in pivot_df.columns:
            pivot_df[variable] = pd.NA

    return pivot_df.sort_values(["ShipName", "EndDateTimeGMT"], ascending=[True, False]).reset_index(drop=True)


# =============================================================================
# Display/export helpers
# =============================================================================


def format_display_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    display_df = df.copy()
    for column in ["StartDateTimeGMT", "EndDateTimeGMT"]:
        if column in display_df.columns:
            display_df[column] = pd.to_datetime(display_df[column], errors="coerce").dt.strftime(DISPLAY_DATETIME_FORMAT)
    for column in display_df.columns:
        if column in {"ReportId", "ShipName", "ReportType", "StartDateTimeGMT", "EndDateTimeGMT", "StateName"}:
            continue
        values = pd.to_numeric(display_df[column], errors="coerce")
        if values.notna().any():
            display_df[column] = values.map(lambda value: "-" if pd.isna(value) else f"{value:,.3f}")
    return display_df.fillna("-")


METRIC_ICON_SVGS = {
    "table_eye": '<svg viewBox="0 0 24 24" aria-hidden="true"><rect x="3" y="5" width="18" height="14" rx="1.8"></rect><path d="M3 10h18M8 5v14M16 5v14"></path><path d="M8.4 12.3c1.1-1.3 2.3-1.9 3.6-1.9s2.5.6 3.6 1.9c-1.1 1.3-2.3 1.9-3.6 1.9s-2.5-.6-3.6-1.9Z"></path><circle cx="12" cy="12.3" r="1.1"></circle></svg>',
    "checked_columns": '<svg viewBox="0 0 24 24" aria-hidden="true"><rect x="4" y="4" width="16" height="16" rx="2"></rect><path d="M9 4v16M15 4v16M7.2 12.2l1.9 1.9 3.8-4.4"></path></svg>',
    "database_rows": '<svg viewBox="0 0 24 24" aria-hidden="true"><ellipse cx="12" cy="5" rx="8" ry="3"></ellipse><path d="M4 5v6c0 1.7 3.6 3 8 3s8-1.3 8-3V5"></path><path d="M4 11v6c0 1.7 3.6 3 8 3s8-1.3 8-3v-6"></path><path d="M8 9h8M8 15h8"></path></svg>',
    "columns_plus": '<svg viewBox="0 0 24 24" aria-hidden="true"><rect x="3" y="4" width="13" height="16" rx="2"></rect><path d="M7.3 4v16M11.7 4v16M18 9v8M14 13h8"></path></svg>',
    "average": '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M4 19V5"></path><path d="M7 17V9M12 17V6M17 17v-5"></path><path d="M4 12h16"></path></svg>',
    "total": '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M17.5 5H7l6 7-6 7h10.5"></path></svg>',
    "numeric": '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M4 8h3v8M4 16h6M13 8h4l-4 8h4M20 8v8"></path></svg>',
    "missing": '<svg viewBox="0 0 24 24" aria-hidden="true"><rect x="4" y="4" width="16" height="16" rx="2" stroke-dasharray="3 2"></rect><path d="M10 9.5a2.2 2.2 0 1 1 3.3 1.9c-.8.5-1.3 1-1.3 2.1"></path><path d="M12 17h.01"></path></svg>',
    # Backwards-compatible names used by older cards.
    "table": '<svg viewBox="0 0 24 24" aria-hidden="true"><rect x="3" y="3" width="18" height="18" rx="1.8"></rect><path d="M3 9h18M3 15h18M9 3v18M15 3v18"></path></svg>',
    "list": '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M8 6h13M8 12h13M8 18h9"></path><path d="M3.5 6h.01M3.5 12h.01M3.5 18h.01"></path></svg>',
    "database": '<svg viewBox="0 0 24 24" aria-hidden="true"><ellipse cx="12" cy="5" rx="8" ry="3"></ellipse><path d="M4 5v6c0 1.7 3.6 3 8 3s8-1.3 8-3V5"></path><path d="M4 11v6c0 1.7 3.6 3 8 3s8-1.3 8-3v-6"></path></svg>',
    "nodes": '<svg viewBox="0 0 24 24" aria-hidden="true"><circle cx="12" cy="4" r="2.5"></circle><circle cx="5" cy="19" r="2.5"></circle><circle cx="19" cy="19" r="2.5"></circle><path d="M10.8 6.2 6.2 16.8M13.2 6.2l4.6 10.6M7.5 19h9"></path></svg>',
}

def render_metric_cards(cards: list[tuple[str, str, str]]) -> None:
    card_html = []
    for label, value, icon_name in cards:
        icon_svg = METRIC_ICON_SVGS.get(icon_name, METRIC_ICON_SVGS["table"])
        card_html.append(
            f'<div class="atlas-metric-card"><div class="atlas-metric-icon">{icon_svg}</div>'
            f'<div><div class="atlas-metric-label">{escape(label)}</div>'
            f'<div class="atlas-metric-value">{escape(value)}</div></div></div>'
        )
    st.markdown(f'<div class="atlas-metric-grid">{"".join(card_html)}</div>', unsafe_allow_html=True)


def render_preview_table(df: pd.DataFrame, row_limit: int = TABLE_PREVIEW_ROW_LIMIT) -> None:
    preview_df = format_display_dataframe(df.head(row_limit))
    if preview_df.empty:
        st.markdown('<div class="atlas-table-empty">No rows to display.</div>', unsafe_allow_html=True)
        return

    columns = [str(column) for column in preview_df.columns]
    header_html = "".join(f"<th>{escape(column)}</th>" for column in columns)
    rows_html: list[str] = []
    for row in preview_df.itertuples(index=False, name=None):
        cell_html = "".join(("<td>" + escape(str(value)) + "</td>") for value in row)
        rows_html.append(f"<tr>{cell_html}</tr>")

    table_html = (
        '<div class="atlas-table-frame"><div class="atlas-table-scroll">'
        f'<table class="atlas-preview-table"><thead><tr>{header_html}</tr></thead>'
        f'<tbody>{"".join(rows_html)}</tbody></table></div></div>'
    )
    st.markdown(table_html, unsafe_allow_html=True)


def make_unique_excel_columns(columns: list[Any]) -> list[str]:
    """Return Excel-table-safe, unique column names while preserving readable labels."""
    safe_columns: list[str] = []
    seen: dict[str, int] = {}

    for position, column in enumerate(columns, start=1):
        label = str(column).strip() if column is not None else ""
        label = re.sub(r"[\x00-\x1f]", "", label)
        if not label or label.lower() in {"nan", "nat", "none"}:
            label = f"Column {position}"
        label = label[:240]

        key = label.casefold()
        count = seen.get(key, 0)
        if count:
            suffix = f"_{count + 1}"
            label = f"{label[:240 - len(suffix)]}{suffix}"
            key = label.casefold()
        seen[key] = count + 1
        safe_columns.append(label)

    return safe_columns


def make_excel_safe_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    safe_df = df.copy()
    safe_df.columns = make_unique_excel_columns(list(safe_df.columns))
    safe_df = safe_df.replace([float("inf"), float("-inf")], pd.NA)
    for column in safe_df.columns:
        if pd.api.types.is_datetime64_any_dtype(safe_df[column]):
            safe_df[column] = pd.to_datetime(safe_df[column], errors="coerce").dt.tz_localize(None)
    return safe_df


def add_excel_table(worksheet: Any, table_name: str) -> None:
    if worksheet.max_row < 2 or worksheet.max_column < 1:
        return

    headers = [worksheet.cell(row=1, column=col).value for col in range(1, worksheet.max_column + 1)]
    if any(header is None or str(header).strip() == "" for header in headers):
        return
    if len({str(header).casefold() for header in headers}) != len(headers):
        return

    table_ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        # Built-in table style kept neutral; explicit teal formatting below
        # gives the export a stable AtlasFlow look in Excel.
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def apply_teal_excel_table_format(worksheet: Any) -> None:
    """Apply AtlasFlow teal styling to exported Excel tables."""
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return

    header_fill = PatternFill(fill_type="solid", fgColor="006B68")
    even_fill = PatternFill(fill_type="solid", fgColor="EAF7F5")
    odd_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    border_side = Side(style="thin", color="B7DCD8")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border

    for row_number in range(2, worksheet.max_row + 1):
        fill = even_fill if row_number % 2 == 0 else odd_fill
        for cell in worksheet[row_number]:
            cell.fill = fill
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center", wrap_text=False)


def autofit_excel_columns(worksheet: Any, max_width: int = 48) -> None:
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), max_width)


def write_table_sheet(writer: Any, df: pd.DataFrame, sheet_name: str, table_name: str) -> None:
    safe_df = make_excel_safe_dataframe(df)
    safe_df.to_excel(writer, index=False, sheet_name=sheet_name)
    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes = "A2"
    autofit_excel_columns(worksheet)
    add_excel_table(worksheet, table_name)
    apply_teal_excel_table_format(worksheet)


def to_excel_bytes(clean_df: pd.DataFrame, pivot_analysis_df: pd.DataFrame | None = None) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        write_table_sheet(writer, clean_df, "Clean Dataset", "AtlasFlowCleanDataset")

        if pivot_analysis_df is not None and not pivot_analysis_df.empty:
            write_table_sheet(writer, pivot_analysis_df, "Summary Analysis", "AtlasFlowSummaryAnalysis")

    return output.getvalue()


def to_displayed_table_excel_bytes(display_df: pd.DataFrame, sheet_name: str = "Displayed Table") -> bytes:
    """Export only the table currently visible to the user.

    This avoids preparing multiple hidden sheets during a normal tab export and keeps
    memory usage lower on Streamlit Cloud.
    """
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        write_table_sheet(writer, display_df, sheet_name[:31], "AtlasFlowDisplayedTable")
    return output.getvalue()


def flatten_pivot_columns(df: pd.DataFrame) -> pd.DataFrame:
    flat_df = df.copy()
    if isinstance(flat_df.columns, pd.MultiIndex):
        flat_df.columns = [
            " | ".join(str(part) for part in column if str(part) not in {"", "nan", "NaT"})
            for column in flat_df.columns.to_flat_index()
        ]
    else:
        flat_df.columns = [str(column) for column in flat_df.columns]
    return flat_df


def build_summary_analysis(
    clean_df: pd.DataFrame,
    group_fields: list[str],
    value_fields: list[str],
    aggregation: str,
) -> pd.DataFrame:
    if clean_df.empty or not group_fields or not value_fields:
        return pd.DataFrame()

    valid_group_fields = [column for column in group_fields if column in clean_df.columns]
    valid_value_fields = [column for column in value_fields if column in clean_df.columns]
    if not valid_group_fields or not valid_value_fields:
        return pd.DataFrame()

    source_df = clean_df.copy()
    for column in valid_value_fields:
        source_df[column] = pd.to_numeric(source_df[column], errors="coerce")

    aggregation_map = {
        "Average": "mean",
        "Sum": "sum",
        "Count": "count",
        "Minimum": "min",
        "Maximum": "max",
        "Median": "median",
    }
    aggfunc = aggregation_map.get(aggregation, "mean")

    summary_df = (
        source_df
        .groupby(valid_group_fields, dropna=False, as_index=False)[valid_value_fields]
        .agg(aggfunc)
    )

    if aggregation != "Count":
        for column in valid_value_fields:
            summary_df[column] = pd.to_numeric(summary_df[column], errors="coerce").round(3)

    rename_map = {column: f"{aggregation} {column}" for column in valid_value_fields}
    return summary_df.rename(columns=rename_map)


def numeric_column_options(df: pd.DataFrame) -> list[str]:
    options: list[str] = []
    for column in df.columns:
        values = pd.to_numeric(df[column], errors="coerce")
        if values.notna().any():
            options.append(column)
    return options


def filter_digest(column: str) -> str:
    return sha256(column.encode("utf-8")).hexdigest()[:10]


def parse_optional_float(value: str) -> tuple[float | None, bool]:
    text = str(value or "").strip()
    if not text:
        return None, True
    normalized = text.replace(" ", "").replace(",", "")
    try:
        return float(normalized), True
    except ValueError:
        return None, False


def parse_optional_date(value: str) -> tuple[pd.Timestamp | None, bool]:
    text = str(value or "").strip()
    if not text:
        return None, True
    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce", utc=True)
    if pd.isna(parsed):
        return None, False
    return parsed, True


def unique_display_values(series: pd.Series, limit: int = 500) -> list[str]:
    values = series.astype("string").fillna("(Blank)").drop_duplicates().tolist()
    values = sorted(values, key=lambda value: str(value).casefold())
    return values[:limit]


def is_numeric_like(series: pd.Series) -> bool:
    values = pd.to_numeric(series, errors="coerce")
    return values.notna().any()


def render_column_filters(df: pd.DataFrame, filter_columns: list[str]) -> list[dict[str, Any]]:
    specs: list[dict[str, Any]] = []
    for column in filter_columns:
        if column not in df.columns:
            continue

        st.caption(f"Filter: {column}")
        series = df[column]
        digest = filter_digest(column)

        if pd.api.types.is_datetime64_any_dtype(series):
            left, right = st.columns(2)
            from_text = left.text_input("From", key=f"atlas_filter_{digest}_from", placeholder="dd/mm/yyyy")
            to_text = right.text_input("To", key=f"atlas_filter_{digest}_to", placeholder="dd/mm/yyyy")
            from_value, from_ok = parse_optional_date(from_text)
            to_value, to_ok = parse_optional_date(to_text)
            if not from_ok or not to_ok:
                st.warning(f"{column}: enter dates as dd/mm/yyyy or yyyy-mm-dd.")
            specs.append({"column": column, "kind": "datetime", "from": from_value, "to": to_value})
            continue

        if is_numeric_like(series):
            values = pd.to_numeric(series, errors="coerce").dropna()
            if not values.empty:
                st.caption(f"Loaded range: {values.min():,.3f} to {values.max():,.3f}")
            left, right = st.columns(2)
            min_text = left.text_input("Min", key=f"atlas_filter_{digest}_min", placeholder="no minimum")
            max_text = right.text_input("Max", key=f"atlas_filter_{digest}_max", placeholder="no maximum")
            minimum, min_ok = parse_optional_float(min_text)
            maximum, max_ok = parse_optional_float(max_text)
            if not min_ok or not max_ok:
                st.warning(f"{column}: enter numeric Min/Max values only.")
            if minimum is not None and maximum is not None and minimum > maximum:
                minimum, maximum = maximum, minimum
            specs.append({"column": column, "kind": "numeric", "min": minimum, "max": maximum})
            continue

        values_key = f"atlas_filter_{digest}_values"
        selected_values = st.multiselect(
            "Values",
            options=unique_display_values(series),
            key=values_key,
            help="Leave blank to include all values for this column.",
        )
        specs.append({"column": column, "kind": "categorical", "values": selected_values})

    return specs


def apply_column_filters(df: pd.DataFrame, specs: list[dict[str, Any]]) -> pd.DataFrame:
    filtered = df.copy()
    for spec in specs:
        column = spec.get("column")
        if column not in filtered.columns:
            continue

        kind = spec.get("kind")
        if kind == "numeric":
            values = pd.to_numeric(filtered[column], errors="coerce")
            minimum = spec.get("min")
            maximum = spec.get("max")
            if minimum is not None:
                filtered = filtered[values >= minimum]
                values = pd.to_numeric(filtered[column], errors="coerce")
            if maximum is not None:
                filtered = filtered[values <= maximum]

        elif kind == "datetime":
            values = pd.to_datetime(filtered[column], errors="coerce", utc=True)
            from_value = spec.get("from")
            to_value = spec.get("to")
            if from_value is not None:
                filtered = filtered[values >= from_value]
                values = pd.to_datetime(filtered[column], errors="coerce", utc=True)
            if to_value is not None:
                filtered = filtered[values < (to_value + pd.Timedelta(days=1))]

        elif kind == "categorical":
            selected_values = spec.get("values") or []
            if selected_values:
                values = filtered[column].astype("string").fillna("(Blank)")
                filtered = filtered[values.isin(selected_values)]

    return filtered



# =============================================================================
# Descriptive statistics helpers
# =============================================================================


def dataframe_numeric_options(df: pd.DataFrame) -> list[str]:
    return [column for column in df.columns if pd.to_numeric(df[column], errors="coerce").notna().any()]


def dataframe_categorical_options(df: pd.DataFrame) -> list[str]:
    options: list[str] = []
    for column in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[column]):
            continue
        numeric_values = pd.to_numeric(df[column], errors="coerce")
        if numeric_values.notna().any():
            continue
        if df[column].astype("string").dropna().nunique() <= 250:
            options.append(column)
    return options


def detect_analysis_datetime_column(df: pd.DataFrame) -> str | None:
    preferred_columns = ["StartDateTimeGMT", "EndDateTimeGMT", "DateTime", "ReportDateTime", "Timestamp"]
    for column in preferred_columns:
        if column in df.columns and pd.to_datetime(df[column], errors="coerce", utc=True).notna().any():
            return column
    for column in df.columns:
        lower = str(column).lower()
        if ("date" in lower or "time" in lower) and pd.to_datetime(df[column], errors="coerce", utc=True).notna().any():
            return column
    return None


def build_descriptive_statistics(df: pd.DataFrame, metric_column: str) -> pd.DataFrame:
    values = pd.to_numeric(df[metric_column], errors="coerce")
    clean_values = values.dropna()
    if clean_values.empty:
        return pd.DataFrame()
    rows = [
        ("Rows", len(df)),
        ("Numeric values", int(clean_values.count())),
        ("Missing values", int(values.isna().sum())),
        ("Sum", clean_values.sum()),
        ("Mean", clean_values.mean()),
        ("Median", clean_values.median()),
        ("Std dev", clean_values.std()),
        ("Minimum", clean_values.min()),
        ("P10", clean_values.quantile(0.10)),
        ("P25", clean_values.quantile(0.25)),
        ("P75", clean_values.quantile(0.75)),
        ("P90", clean_values.quantile(0.90)),
        ("Maximum", clean_values.max()),
    ]
    return pd.DataFrame(rows, columns=["Statistic", "Value"])


def build_grouped_descriptive_statistics(df: pd.DataFrame, metric_column: str, group_column: str) -> pd.DataFrame:
    if group_column not in df.columns or metric_column not in df.columns:
        return pd.DataFrame()
    source = df[[group_column, metric_column]].copy()
    source[metric_column] = pd.to_numeric(source[metric_column], errors="coerce")
    source = source[source[metric_column].notna()].copy()
    if source.empty:
        return pd.DataFrame()
    grouped = (
        source
        .groupby(group_column, dropna=False)[metric_column]
        .agg(Count="count", Sum="sum", Mean="mean", Median="median", Minimum="min", Maximum="max")
        .reset_index()
        .sort_values("Sum", ascending=False)
    )
    for column in ["Sum", "Mean", "Median", "Minimum", "Maximum"]:
        grouped[column] = pd.to_numeric(grouped[column], errors="coerce").round(3)
    return grouped


def build_monthly_trend(df: pd.DataFrame, metric_column: str, datetime_column: str) -> pd.DataFrame:
    source = df[[datetime_column, metric_column]].copy()
    source[datetime_column] = pd.to_datetime(source[datetime_column], errors="coerce", utc=True)
    source[metric_column] = pd.to_numeric(source[metric_column], errors="coerce")
    source = source[source[datetime_column].notna() & source[metric_column].notna()].copy()
    if source.empty:
        return pd.DataFrame()
    source["Month"] = source[datetime_column].dt.to_period("M").astype(str)
    trend = (
        source
        .groupby("Month", as_index=False)[metric_column]
        .agg(Count="count", Sum="sum", Mean="mean", Median="median")
        .sort_values("Month")
    )
    for column in ["Sum", "Mean", "Median"]:
        trend[column] = pd.to_numeric(trend[column], errors="coerce").round(3)
    return trend


def render_descriptive_statistics_tab(
    custom_df: pd.DataFrame,
    reportpivots_df: pd.DataFrame,
    shippivots_df: pd.DataFrame,
) -> None:
    st.markdown('<div class="section-title">Descriptive Statistics</div>', unsafe_allow_html=True)
    st.caption("Analyze exactly the same filtered/export-ready tables from each source. No extra KPI logic is applied here.")

    sources = {
        "Custom Analytics": custom_df,
        "Noon & Manual Reports": reportpivots_df,
        "15-Minute Operations": shippivots_df,
    }
    available_sources = [label for label, df in sources.items() if isinstance(df, pd.DataFrame) and not df.empty]
    if not available_sources:
        st.info("No filtered source table is available for descriptive statistics yet.")
        return

    selected_source = st.selectbox("Source table", options=available_sources, key="atlas_descriptive_source")
    analysis_df = sources[selected_source].copy()
    numeric_options = dataframe_numeric_options(analysis_df)
    if not numeric_options:
        st.info("The selected source table has no numeric columns to analyze.")
        return

    metric_column = st.selectbox("Metric to analyze", options=numeric_options, key="atlas_descriptive_metric")
    group_options = ["None"] + dataframe_categorical_options(analysis_df)
    default_group_index = group_options.index("ShipName") if "ShipName" in group_options else 0
    group_column = st.selectbox("Optional group by", options=group_options, index=default_group_index, key="atlas_descriptive_group")

    stats_df = build_descriptive_statistics(analysis_df, metric_column)
    if stats_df.empty:
        st.info("No numeric values were found for the selected metric.")
        return

    values = pd.to_numeric(analysis_df[metric_column], errors="coerce")
    render_metric_cards(
        [
            ("Numeric Values", f"{values.notna().sum():,}", "numeric"),
            ("Total", f"{values.sum(skipna=True):,.3f}", "total"),
            ("Average", f"{values.mean(skipna=True):,.3f}", "average"),
            ("Missing", f"{values.isna().sum():,}", "missing"),
        ]
    )

    st.markdown('<div class="section-title">Overall statistics</div>', unsafe_allow_html=True)
    st.dataframe(format_display_dataframe(stats_df), use_container_width=True, hide_index=True)

    if group_column != "None":
        grouped_df = build_grouped_descriptive_statistics(analysis_df, metric_column, group_column)
        if not grouped_df.empty:
            st.markdown('<div class="section-title">Grouped statistics</div>', unsafe_allow_html=True)
            st.dataframe(format_display_dataframe(grouped_df.head(100)), use_container_width=True, hide_index=True)

    datetime_column = detect_analysis_datetime_column(analysis_df)
    if datetime_column:
        trend_df = build_monthly_trend(analysis_df, metric_column, datetime_column)
        if not trend_df.empty:
            st.markdown('<div class="section-title">Monthly trend</div>', unsafe_allow_html=True)
            st.dataframe(format_display_dataframe(trend_df), use_container_width=True, hide_index=True)
            chart_df = trend_df.set_index("Month")[["Sum", "Mean"]]
            st.line_chart(chart_df)

    outlier_values = values.dropna()
    if len(outlier_values) >= 4:
        q1 = outlier_values.quantile(0.25)
        q3 = outlier_values.quantile(0.75)
        iqr = q3 - q1
        lower = q1 - 1.5 * iqr
        upper = q3 + 1.5 * iqr
        outliers = analysis_df[(values < lower) | (values > upper)].copy()
        if not outliers.empty:
            st.markdown('<div class="section-title">Potential outlier rows</div>', unsafe_allow_html=True)
            display_cols = [column for column in ["ShipName", "ReportType", "StartDateTimeGMT", "DateTime", metric_column] if column in outliers.columns]
            if not display_cols:
                display_cols = list(outliers.columns[: min(8, len(outliers.columns))])
            st.dataframe(format_display_dataframe(outliers[display_cols].head(100)), use_container_width=True, hide_index=True)


# =============================================================================
# Sidebar/session helpers
# =============================================================================


def request_signature(username: str, auth_method: str, start_date: date) -> dict[str, Any]:
    return {
        "endpoint": ODATA_ENDPOINT,
        "username_hash": sha256(username.encode("utf-8")).hexdigest()[:12],
        "auth_method": auth_method.lower(),
        "start_date": start_date.isoformat(),
    }


def raw_data_covers_request(
    loaded_signature: dict[str, Any] | None,
    metadata: dict[str, Any] | None,
    requested_signature: dict[str, Any],
    requested_start_date: date,
) -> bool:
    if not loaded_signature or not metadata:
        return False
    for key in ["endpoint", "username_hash", "auth_method"]:
        if loaded_signature.get(key) != requested_signature.get(key):
            return False
    loaded_start_text = metadata.get("loaded_start_date") or loaded_signature.get("start_date")
    try:
        loaded_start_date = date.fromisoformat(str(loaded_start_text))
    except ValueError:
        return False
    return loaded_start_date <= requested_start_date


def get_loaded_state() -> tuple[pd.DataFrame | None, pd.DataFrame | None, dict[str, Any] | None]:
    return (
        st.session_state.get("loaded_raw_df"),
        st.session_state.get("loaded_long_df"),
        st.session_state.get("loaded_metadata"),
    )


def set_loaded_raw_state(raw_df: pd.DataFrame, metadata: dict[str, Any], signature: dict[str, Any]) -> None:
    metadata = metadata.copy()
    metadata.setdefault("loaded_at_utc", datetime.now(timezone.utc).strftime("%d-%m-%Y %H:%M:%S UTC"))
    metadata.setdefault("loaded_at_local", local_time_label())
    metadata["loaded_start_date"] = signature["start_date"]
    st.session_state["loaded_raw_df"] = raw_df
    st.session_state["loaded_metadata"] = metadata
    st.session_state["loaded_request_signature"] = signature
    st.session_state.pop("loaded_long_df", None)
    st.session_state.pop("loaded_prepare_signature", None)


def save_raw_snapshot(raw_df: pd.DataFrame, metadata: dict[str, Any], signature: dict[str, Any]) -> None:
    """Persist the latest successful raw API pull as a local fallback after app restarts."""
    try:
        SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
        raw_df.to_parquet(RAW_SNAPSHOT_FILE, index=False)
        snapshot_payload = {
            "metadata": metadata,
            "signature": signature,
            "saved_at_utc": datetime.now(timezone.utc).strftime("%d-%m-%Y %H:%M:%S UTC"),
        }
        METADATA_SNAPSHOT_FILE.write_text(json.dumps(snapshot_payload, indent=2, default=str), encoding="utf-8")
    except Exception:
        # Snapshot persistence is a speed fallback only; never break the app if it fails.
        return


def load_raw_snapshot(
    requested_signature: dict[str, Any],
    requested_start_date: date,
) -> tuple[pd.DataFrame, dict[str, Any], dict[str, Any]] | None:
    """Load the latest local raw-data snapshot if it covers the current API request."""
    try:
        if not RAW_SNAPSHOT_FILE.is_file() or not METADATA_SNAPSHOT_FILE.is_file():
            return None
        snapshot_payload = json.loads(METADATA_SNAPSHOT_FILE.read_text(encoding="utf-8"))
        metadata = snapshot_payload.get("metadata") or {}
        signature = snapshot_payload.get("signature") or {}
        if not raw_data_covers_request(signature, metadata, requested_signature, requested_start_date):
            return None
        raw_df = pd.read_parquet(RAW_SNAPSHOT_FILE)
        if not isinstance(raw_df, pd.DataFrame) or raw_df.empty:
            return None
        metadata = metadata.copy()
        metadata["loaded_from_snapshot"] = True
        metadata.setdefault("snapshot_saved_at_utc", snapshot_payload.get("saved_at_utc", "-"))
        return raw_df, metadata, signature
    except Exception:
        return None


def set_loaded_long_state(long_df: pd.DataFrame, signature: dict[str, Any]) -> None:
    st.session_state["loaded_long_df"] = long_df
    st.session_state["loaded_prepare_signature"] = signature


def selected_vessel_controls() -> tuple[str, list[str]]:
    group_options = ["Single vessel", "All fleets"] + list(VESSEL_GROUPS.keys())
    selected_group = st.sidebar.selectbox("Fleet group", options=group_options, key="atlas_fleet_group")

    if selected_group == "Single vessel":
        vessel = st.sidebar.selectbox("Vessel to include", options=VESSEL_OPTIONS, key="atlas_single_vessel")
        st.session_state["atlas_last_fleet_group"] = selected_group
        return selected_group, [vessel]

    if selected_group == "All fleets":
        group_vessels = VESSEL_OPTIONS
    else:
        group_vessels = VESSEL_GROUPS[selected_group]

    vessel_key = "atlas_selected_vessels"
    last_group_key = "atlas_last_fleet_group"
    previous_group = st.session_state.get(last_group_key)

    # Streamlit keeps multiselect state after the first render, so changing from
    # Fleet 1 to All fleets could otherwise keep only the old Fleet 1 vessels.
    # Reset to the full new group whenever the fleet-group selector changes.
    if previous_group != selected_group:
        st.session_state[vessel_key] = list(group_vessels)
        st.session_state[last_group_key] = selected_group

    previous_vessels = st.session_state.get(vessel_key, group_vessels)
    if not isinstance(previous_vessels, list):
        previous_vessels = group_vessels
    valid_default_vessels = [vessel for vessel in previous_vessels if vessel in group_vessels]
    if not valid_default_vessels:
        valid_default_vessels = list(group_vessels)
    if valid_default_vessels != previous_vessels:
        st.session_state[vessel_key] = valid_default_vessels

    vessels = st.sidebar.multiselect(
        "Vessels to include",
        options=group_vessels,
        default=valid_default_vessels,
        key=vessel_key,
        help="This controls the displayed datasets only. The API data is loaded broadly.",
    )

    if not vessels:
        st.sidebar.caption("No vessels selected manually, so all vessels in this fleet group are included.")
        vessels = group_vessels

    return selected_group, list(vessels)

def sidebar_refresh_control() -> bool:
    refresh_requested = st.sidebar.button("Refresh all APIs", use_container_width=False)
    if refresh_requested:
        st.session_state["confirm_api_refresh"] = True

    refresh = False
    if st.session_state.get("confirm_api_refresh"):
        metadata = st.session_state.get("loaded_metadata") or {}
        last_load = metadata.get("loaded_at_local") or metadata.get("loaded_at_utc") or "-"
        last_load_display = str(last_load).replace(" EEST", "").replace(" EET", "")
        st.sidebar.warning(
            "Refresh will call all AtlasFlow APIs and may take a while.\n\n"
            f"Last updated data was on: {last_load_display} LT"
        )
        col1, col2 = st.sidebar.columns(2)
        if col1.button("Confirm"):
            refresh = True
            st.session_state["confirm_api_refresh"] = False
        if col2.button("Cancel"):
            st.session_state["confirm_api_refresh"] = False
            st.rerun()
    return refresh


def add_months(base_date: date, months: int) -> date:
    month_index = base_date.month - 1 + months
    year = base_date.year + month_index // 12
    month = month_index % 12 + 1
    days_in_month = [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    day = min(base_date.day, days_in_month[month - 1])
    return date(year, month, day)


def first_day_of_month(value: date) -> date:
    return date(value.year, value.month, 1)


def last_day_previous_month(value: date) -> date:
    return first_day_of_month(value) - timedelta(days=1)


def clamp_period(start_value: date, end_value: date, min_date: date, max_date: date) -> tuple[date, date]:
    start_value = max(start_value, min_date)
    end_value = min(end_value, max_date)
    if start_value > end_value:
        return min_date, max_date
    return start_value, end_value


def dynamic_period_dates(period_label: str, min_date: date, max_date: date) -> tuple[date, date]:
    anchor_date = min(max_date, date.today())

    if period_label == "YTD":
        start_value = date(anchor_date.year, 1, 1)
        end_value = anchor_date
    elif period_label == "Previous month":
        end_value = last_day_previous_month(anchor_date)
        start_value = first_day_of_month(end_value)
    elif period_label == "Year to previous month":
        end_value = last_day_previous_month(anchor_date)
        start_value = date(end_value.year, 1, 1)
    elif period_label == "Previous 3 months":
        end_value = anchor_date
        start_value = add_months(anchor_date, -3) + timedelta(days=1)
    elif period_label == "Previous 6 months":
        end_value = anchor_date
        start_value = add_months(anchor_date, -6) + timedelta(days=1)
    elif period_label == "Previous 12 months":
        end_value = anchor_date
        start_value = add_months(anchor_date, -12) + timedelta(days=1)
    else:
        start_value = min_date
        end_value = max_date

    return clamp_period(start_value, end_value, min_date, max_date)


def render_date_slicer(df: pd.DataFrame) -> tuple[date, date]:
    min_date, max_date = dataframe_date_window(df)
    st.sidebar.markdown("### Period")
    if min_date >= max_date:
        st.sidebar.caption(f"Available data period: {min_date.strftime('%d/%m/%Y')}")
        return min_date, max_date

    period_mode = st.sidebar.selectbox(
        "Period preset",
        options=[
            "Custom range",
            "YTD",
            "Previous month",
            "Year to previous month",
            "Previous 3 months",
            "Previous 6 months",
            "Previous 12 months",
            "Full available period",
        ],
        index=0,
        key="atlas_period_preset",
        help="Use a dynamic preset or choose Custom range to control the period manually with the slider.",
    )

    if period_mode == "Custom range":
        selected_start, selected_end = st.sidebar.slider(
            "Report period",
            min_value=min_date,
            max_value=max_date,
            value=(min_date, max_date),
            format="DD/MM/YYYY",
            key="atlas_period_slicer",
        )
    else:
        selected_start, selected_end = dynamic_period_dates(period_mode, min_date, max_date)
        st.sidebar.caption(
            f"Selected period: {selected_start.strftime('%d/%m/%Y')} to {selected_end.strftime('%d/%m/%Y')}"
        )

    return selected_start, selected_end




# =============================================================================
# Streamlit Cloud-safe snapshot refresh helpers
# =============================================================================


def normalize_snapshot_values(df: pd.DataFrame) -> pd.DataFrame:
    """Store raw API values as nullable strings to keep Parquet schemas stable page by page."""
    if df.empty:
        return df.copy()
    safe_df = df.copy()
    for column in safe_df.columns:
        safe_df[column] = safe_df[column].astype("string")
    return safe_df


def write_parquet_pages(page_frames: list[pd.DataFrame], output_file: Path) -> int:
    """Write already-normalized page frames into one Parquet file with a stable schema."""
    output_file.parent.mkdir(parents=True, exist_ok=True)
    writer = None
    row_count = 0
    try:
        for frame in page_frames:
            if frame.empty:
                continue
            normalized = normalize_snapshot_values(frame)
            table = pa.Table.from_pandas(normalized, preserve_index=False)
            if writer is None:
                writer = pq.ParquetWriter(output_file, table.schema, compression="zstd")
            writer.write_table(table)
            row_count += len(normalized)
    finally:
        if writer is not None:
            writer.close()
    return row_count


def fetch_report_data_to_snapshot(
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
) -> dict[str, Any]:
    """Fetch ReportData page-by-page and write to Parquet without keeping the full dataset in memory."""
    started_at = time.perf_counter()
    next_url = build_odata_url(start_date)
    first_url = next_url
    seen_urls: set[str] = set()
    pages = 0
    total_bytes = 0
    scanned_rows = 0
    kept_rows_total = 0
    # Use a unique temporary file per warmup request. Streamlit Cloud can run
    # multiple warmup/browser sessions at the same time; a fixed .tmp.parquet
    # name can be deleted by another session before this request reaches replace().
    tmp_file = RAW_SNAPSHOT_FILE.with_name(
        f"{RAW_SNAPSHOT_FILE.stem}.{os.getpid()}.{int(time.time() * 1000)}.tmp.parquet"
    )
    tmp_file.parent.mkdir(parents=True, exist_ok=True)
    if tmp_file.exists():
        tmp_file.unlink()

    auth = request_auth(username, password, auth_method)
    headers = request_headers(token, auth_method)
    writer = None

    try:
        with requests.Session() as session:
            session.headers.update(headers)
            for _ in range(MAX_ODATA_PAGES):
                if next_url in seen_urls:
                    break
                seen_urls.add(next_url)

                response = session.get(next_url, auth=auth, timeout=90)
                total_bytes += len(response.content)
                response.raise_for_status()
                pages += 1

                page_rows, next_link = extract_odata_page(response.json())
                scanned_rows += len(page_rows)
                compact_rows = compact_odata_rows(page_rows)
                if compact_rows:
                    page_df = normalize_snapshot_values(rows_to_dataframe(compact_rows))
                    table = pa.Table.from_pandas(page_df, preserve_index=False)
                    if writer is None:
                        writer = pq.ParquetWriter(tmp_file, table.schema, compression="zstd")
                    writer.write_table(table)
                    kept_rows_total += len(page_df)
                    del page_df, table

                del page_rows, compact_rows
                gc.collect()

                if not next_link:
                    break
                next_url = urljoin(next_url, next_link)
    finally:
        if writer is not None:
            writer.close()

    RAW_SNAPSHOT_FILE.parent.mkdir(parents=True, exist_ok=True)

    if kept_rows_total == 0:
        empty_df = normalize_snapshot_values(pd.DataFrame(columns=SOURCE_COLUMNS))
        empty_table = pa.Table.from_pandas(empty_df, preserve_index=False)
        pq.write_table(empty_table, tmp_file, compression="zstd")
        del empty_df, empty_table

    if not tmp_file.exists():
        raise FileNotFoundError(f"ReportData snapshot temporary file was not created: {tmp_file}")

    tmp_file.replace(RAW_SNAPSHOT_FILE)
    loaded_at_utc = datetime.now(timezone.utc)
    metadata = {
        "loaded_at_utc": loaded_at_utc.strftime("%d-%m-%Y %H:%M:%S UTC"),
        "loaded_at_local": local_time_label(loaded_at_utc),
        "rows": kept_rows_total,
        "kept_rows": kept_rows_total,
        "scanned_rows": scanned_rows,
        "discarded_rows": max(scanned_rows - kept_rows_total, 0),
        "pages": pages,
        "downloaded_mb": round(total_bytes / 1024 / 1024, 2),
        "fetch_seconds": round(time.perf_counter() - started_at, 2),
        "first_url": first_url,
        "hit_page_limit": pages >= MAX_ODATA_PAGES,
        "loaded_start_date": start_date.isoformat(),
        "snapshot_format": "parquet",
        "reportdata_mode": "atlasflow_consumption_oil_stats_whitelist",
        "value_description_whitelist_count": len(REPORTDATA_VALUE_WHITELIST),
    }
    signature = request_signature(username, auth_method, start_date)
    snapshot_payload = {
        "metadata": metadata,
        "signature": signature,
        "saved_at_utc": datetime.now(timezone.utc).strftime("%d-%m-%Y %H:%M:%S UTC"),
    }
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    METADATA_SNAPSHOT_FILE.write_text(json.dumps(snapshot_payload, indent=2, default=str), encoding="utf-8")
    return metadata


def fetch_wide_source_to_snapshot(
    source_key: str,
    username: str,
    password: str,
    token: str,
    auth_method: str,
    start_date: date,
) -> dict[str, Any]:
    """Fetch a wide OData source page-by-page and write to Parquet safely."""
    config = SOURCE_CONFIGS[source_key]
    endpoint = str(config["endpoint"])
    datetime_column = str(config.get("datetime_candidates", ["DateTime"])[0])
    next_url = build_wide_odata_url(endpoint, start_date, datetime_column)
    first_url = next_url
    seen_urls: set[str] = set()
    pages = 0
    total_bytes = 0
    row_count = 0
    all_columns: list[str] = []
    target_file = Path(config["snapshot_file"])
    # Use a unique temporary file per warmup request to avoid cross-session
    # collisions when several source warmups or browser tabs run concurrently.
    tmp_file = target_file.with_name(
        f"{target_file.stem}.{os.getpid()}.{int(time.time() * 1000)}.tmp.parquet"
    )
    tmp_file.parent.mkdir(parents=True, exist_ok=True)
    if tmp_file.exists():
        tmp_file.unlink()

    auth = request_auth(username, password, auth_method)
    headers = request_headers(token, auth_method)
    writer = None
    started_at = time.perf_counter()

    try:
        with requests.Session() as session:
            session.headers.update(headers)
            for _ in range(MAX_ODATA_PAGES):
                if next_url in seen_urls:
                    break
                seen_urls.add(next_url)
                response = session.get(next_url, auth=auth, timeout=90)
                total_bytes += len(response.content)
                response.raise_for_status()
                pages += 1

                page_rows, next_link = extract_odata_page(response.json())
                page_df = pd.DataFrame(page_rows)
                if "__metadata" in page_df.columns:
                    page_df = page_df.drop(columns=["__metadata"])

                if not page_df.empty:
                    if not all_columns:
                        all_columns = list(page_df.columns)
                    else:
                        for column in page_df.columns:
                            if column not in all_columns:
                                all_columns.append(column)
                        for column in all_columns:
                            if column not in page_df.columns:
                                page_df[column] = pd.NA
                        page_df = page_df[all_columns]

                    page_df = normalize_snapshot_values(page_df)
                    table = pa.Table.from_pandas(page_df, preserve_index=False)
                    if writer is None:
                        writer = pq.ParquetWriter(tmp_file, table.schema, compression="zstd")
                    else:
                        # New columns after the first page are rare; keep only first-page schema to avoid schema errors.
                        schema_names = writer.schema.names
                        page_df = page_df[[column for column in schema_names if column in page_df.columns]]
                        for column in schema_names:
                            if column not in page_df.columns:
                                page_df[column] = pd.NA
                        page_df = page_df[schema_names]
                        table = pa.Table.from_pandas(page_df, schema=writer.schema, preserve_index=False)
                    writer.write_table(table)
                    row_count += len(page_df)
                    del page_df, table

                del page_rows
                gc.collect()

                if not next_link:
                    break
                next_url = urljoin(next_url, next_link)
    finally:
        if writer is not None:
            writer.close()

    target_file.parent.mkdir(parents=True, exist_ok=True)

    if row_count == 0:
        # Some wide endpoints can legally return no rows for the selected API window.
        # In that case, create an empty snapshot with discovered columns if possible.
        empty_columns = all_columns if all_columns else ["NoData"]
        empty_df = normalize_snapshot_values(pd.DataFrame(columns=empty_columns))
        empty_df.to_parquet(tmp_file, index=False, compression="zstd")
        del empty_df

    if not tmp_file.exists():
        raise FileNotFoundError(f"{config['label']} snapshot temporary file was not created: {tmp_file}")

    # Validate before replacing the previous good snapshot. Do not accept a placeholder
    # NoData file when API rows were written.
    try:
        parquet_columns = pq.ParquetFile(tmp_file).schema.names
    except Exception as exc:
        raise RuntimeError(f"{config['label']} snapshot validation failed before save: {exc}") from exc
    if row_count > 0 and parquet_columns == ["NoData"]:
        raise RuntimeError(f"{config['label']} snapshot validation failed: placeholder NoData file for {row_count:,} rows.")

    if target_file.exists():
        target_file.unlink()
    os.replace(str(tmp_file), str(target_file))
    loaded_at_utc = datetime.now(timezone.utc)
    metadata = {
        "source": config["label"],
        "endpoint": endpoint,
        "loaded_at_utc": loaded_at_utc.strftime("%d-%m-%Y %H:%M:%S UTC"),
        "loaded_at_local": local_time_label(loaded_at_utc),
        "loaded_start_date": start_date.isoformat(),
        "rows": int(row_count),
        "columns": int(len(all_columns)),
        "pages": pages,
        "downloaded_mb": round(total_bytes / 1024 / 1024, 2),
        "fetch_seconds": round(time.perf_counter() - started_at, 2),
        "first_url": first_url,
        "hit_page_limit": pages >= MAX_ODATA_PAGES,
        "snapshot_format": "parquet",
    }
    signature = source_signature(source_key, username, auth_method, start_date)
    payload = {
        "metadata": metadata,
        "signature": signature,
        "saved_at_utc": datetime.now(timezone.utc).strftime("%d-%m-%Y %H:%M:%S UTC"),
    }
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    Path(config["metadata_file"]).write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    return metadata


# =============================================================================
# Warmup
# =============================================================================


def run_warmup_if_requested() -> None:
    if not is_warmup_request():
        return

    apply_custom_css()
    if not warmup_token_is_valid():
        st.error("Invalid or missing warmup token.")
        st.stop()

    username = read_secret("MARORKA_USERNAME")
    password = read_secret("MARORKA_PASSWORD")
    token = read_secret("MARORKA_TOKEN")
    auth_method = read_secret("MARORKA_AUTH_METHOD", "basic")
    start_date = API_FULL_START_DATE
    requested_source = get_query_param("source", "reportdata").strip().lower()

    if auth_method.lower() in {"basic", "digest"} and (not username or not password):
        st.error("Warmup failed: MARORKA_USERNAME and MARORKA_PASSWORD are required.")
        st.stop()

    valid_sources = {"reportdata", "reportpivots", "shippivots"}
    if requested_source == "all":
        st.error(
            "source=all is disabled on Streamlit Cloud to avoid memory limits. "
            "Warm up reportdata, reportpivots, and shippivots one at a time."
        )
        st.stop()
    if requested_source not in valid_sources:
        st.error("Invalid warmup source. Use reportdata, reportpivots, or shippivots.")
        st.stop()

    force_refresh = get_query_param("force", "0") == "1"

    # Do not clear active Streamlit caches before the new API pull succeeds.
    # The snapshot writers already use unique temporary parquet files and only
    # replace the previous good snapshot after the fresh fetch has completed.
    # Clearing cached data after success avoids an empty/slow window for users
    # who are actively browsing while Task Scheduler is warming up ReportData.
    try:
        with st.spinner(f"Warming up AtlasFlow source: {requested_source}..."):
            if requested_source == "reportdata":
                metadata = fetch_report_data_to_snapshot(username, password, token, auth_method, start_date)
            else:
                metadata = fetch_wide_source_to_snapshot(requested_source, username, password, token, auth_method, start_date)

            if force_refresh:
                if requested_source == "reportdata":
                    cached_fetch_report_data.clear()
                    cached_prepare_long_data.clear()
                    build_pivot_table.clear()
                else:
                    cached_fetch_wide_odata_source.clear()
                gc.collect()
    except requests.HTTPError as exc:
        status = exc.response.status_code if exc.response is not None else "unknown"
        st.error(f"Warmup failed: Marorka API request failed with status {status}.")
        st.stop()
    except (MarorkaConfigError, ValueError, requests.RequestException) as exc:
        st.error(f"Warmup failed: {exc}")
        st.stop()

    st.success("Warmup OK.")
    st.write(
        {
            "source": requested_source,
            "last_api_load_local": metadata.get("loaded_at_local", "-"),
            "rows": metadata.get("rows", metadata.get("kept_rows", 0)),
            "pages": metadata.get("pages", 0),
            "downloaded_mb": metadata.get("downloaded_mb", "-"),
            "fetch_seconds": metadata.get("fetch_seconds", "-"),
            "snapshot_format": "parquet",
            "force_refresh": force_refresh,
        }
    )
    st.stop()

# =============================================================================
# Main app
# =============================================================================


def main() -> None:
    run_warmup_if_requested()
    require_dashboard_password()
    apply_custom_css()

    username = read_secret("MARORKA_USERNAME")
    password = read_secret("MARORKA_PASSWORD")
    token = read_secret("MARORKA_TOKEN")
    auth_method = read_secret("MARORKA_AUTH_METHOD", "basic")

    if auth_method.lower() in {"basic", "digest"} and (not username or not password):
        st.info("Add MARORKA_USERNAME and MARORKA_PASSWORD to .streamlit/secrets.toml or Streamlit Cloud Secrets.")
        st.stop()

    api_start_date = API_FULL_START_DATE
    api_end_date = date.today()

    refresh = sidebar_refresh_control()
    selected_group, selected_vessels = selected_vessel_controls()

    raw_signature = request_signature(username, auth_method, api_start_date)
    current_raw_signature = st.session_state.get("loaded_request_signature")
    raw_df, long_df, metadata = get_loaded_state()

    needs_raw_load = (
        refresh
        or raw_df is None
        or metadata is None
        or not raw_data_covers_request(current_raw_signature, metadata, raw_signature, api_start_date)
    )

    if needs_raw_load and not refresh:
        snapshot = load_raw_snapshot(raw_signature, api_start_date)
        if snapshot is not None:
            raw_df, metadata, snapshot_signature = snapshot
            set_loaded_raw_state(raw_df, metadata, snapshot_signature)
            long_df = None
            needs_raw_load = False

    if needs_raw_load:
        st.warning(
            "No usable ReportData snapshot is available yet. "
            "Run the ReportData warmup URL first so AtlasFlow can open from stored Parquet data instead of live-loading the large API in the UI."
        )
        st.code(
            "https://atlas-flow.streamlit.app/?warmup=1&force=1&source=reportdata&token=warmup-atlas-flow",
            language="text",
        )
        st.stop()

    prepare_signature = {**raw_signature, "prepare_version": "atlasflow_dynamic_pivot_v3_oil_stats"}
    current_prepare_signature = st.session_state.get("loaded_prepare_signature")
    raw_df = st.session_state.get("loaded_raw_df")
    long_df = st.session_state.get("loaded_long_df")
    metadata = st.session_state.get("loaded_metadata")

    if raw_df is None or metadata is None:
        st.info("Loading Marorka data automatically. Use Refresh API data to force a new API pull.")
        st.stop()

    if long_df is None or current_prepare_signature != prepare_signature:
        try:
            transform_started_at = time.perf_counter()
            long_df = cached_prepare_long_data(raw_df)
            set_loaded_long_state(long_df, prepare_signature)
            metadata = st.session_state.get("loaded_metadata")
            if isinstance(metadata, dict):
                metadata["prepare_seconds"] = round(time.perf_counter() - transform_started_at, 2)
                metadata["long_rows"] = int(len(long_df))
                metadata["available_variables"] = int(long_df["ValueDescription"].nunique()) if "ValueDescription" in long_df.columns else 0
                st.session_state["loaded_metadata"] = metadata
        except (ValueError, TypeError) as exc:
            st.error(str(exc))
            st.stop()

    if long_df.empty:
        render_header(selected_group, selected_vessels, [])
        render_api_load_caption(metadata)
        st.warning("No Marorka report values were returned for the loaded API window.")
        st.stop()

    selected_start, selected_end = render_date_slicer(long_df)

    reportpivots_df, reportpivots_metadata = load_or_fetch_source(
        "reportpivots", username, password, token, auth_method, api_start_date, refresh
    )
    shippivots_df, shippivots_metadata = load_or_fetch_source(
        "shippivots", username, password, token, auth_method, api_start_date, refresh
    )

    # ReportType is no longer a separate sidebar filter.
    # It is handled together with all other displayed columns inside
    # "Filters for displayed columns" to keep one unified filter section.
    selected_report_types: list[str] = []

    filtered_long_for_options = filter_long_data(
        long_df,
        selected_vessels=selected_vessels,
        selected_report_types=selected_report_types,
        selected_start=selected_start,
        selected_end=selected_end,
    )

    variable_options = sorted(
        set(available_variables(filtered_long_for_options)).union(DERIVED_VARIABLES),
        key=str.casefold,
    )
    st.sidebar.markdown("### Pivot variables")
    previous_selected_variables = st.session_state.get("atlas_selected_variables", [])
    if not isinstance(previous_selected_variables, list):
        previous_selected_variables = []
    valid_default_variables = [
        variable for variable in previous_selected_variables
        if variable in variable_options
    ]
    selected_variables = st.sidebar.multiselect(
        "Variables to include and filter",
        options=variable_options,
        default=valid_default_variables,
        key="atlas_selected_variables",
        help=(
            "Every selected ValueDescription becomes a displayed table column. "
            "The same selected variables are also available below as filters."
        ),
    )

    identity_columns = st.sidebar.multiselect(
        "Pivot rows",
        options=PIVOT_IDENTITY_COLUMNS,
        default=["ShipName"],
        help="Choose the row fields that appear before the selected variable columns.",
    )
    if not identity_columns:
        identity_columns = ["ShipName"]

    render_header(selected_group, selected_vessels, selected_variables)
    render_api_load_caption(metadata)

    if not selected_variables:
        st.info("Select one or more variables from the sidebar to build the AtlasFlow pivot table.")

    pivot_df = build_pivot_table(filtered_long_for_options, tuple(selected_variables))

    filter_column_options = [column for column in [*identity_columns, *selected_variables] if column in pivot_df.columns]
    with st.sidebar.expander("Filters for displayed columns", expanded=False):
        st.caption("Choose columns to filter. Selected variables are already part of the displayed table.")
        previous_filter_columns = st.session_state.get("atlas_columns_to_filter", [])
        if not isinstance(previous_filter_columns, list):
            previous_filter_columns = []
        valid_filter_columns = [column for column in previous_filter_columns if column in filter_column_options]
        if valid_filter_columns != previous_filter_columns:
            st.session_state["atlas_columns_to_filter"] = valid_filter_columns

        columns_to_filter = st.multiselect(
            "Columns to filter",
            options=filter_column_options,
            default=valid_filter_columns,
            key="atlas_columns_to_filter",
        )
        filter_specs = render_column_filters(pivot_df, columns_to_filter)

    filtered_pivot_df = apply_column_filters(pivot_df, filter_specs)

    display_columns = []
    for column in [*identity_columns, *selected_variables]:
        if column in filtered_pivot_df.columns and column not in display_columns:
            display_columns.append(column)
    if not display_columns:
        display_columns = [column for column in DEFAULT_DISPLAY_IDENTITY_COLUMNS if column in filtered_pivot_df.columns]

    output_df = filtered_pivot_df[display_columns].copy()

    tab_table, tab_reportpivots, tab_shippivots, tab_descriptive, tab_export, tab_diagnostics = st.tabs([
        "Custom Analytics",
        "Noon & Manual Reports",
        "15-Minute Operations",
        "Descriptive Statistics",
        "Export Center",
        "API Diagnostics",
    ])

    if metadata.get("hit_page_limit"):
        st.warning(
            "The API refresh reached the page safety limit. The loaded dataset may be incomplete. "
            "Check API Diagnostics before using the export."
        )

    with tab_table:
        st.markdown('<div class="section-title">Custom Analytics Preview & Export</div>', unsafe_allow_html=True)

        summary_builder_columns = [column for column in output_df.columns]
        summary_value_options = numeric_column_options(output_df)

        st.caption(
            "Choose which table you want to preview and export. The visible table below is the same table prepared for Excel."
        )
        preview_mode = st.radio(
            "Preview table",
            options=["Clean Dataset", "Summary Analysis", "Source Data"],
            horizontal=True,
            key="atlas_reportdata_preview_mode",
        )

        if preview_mode == "Summary Analysis":
            st.markdown('<div class="section-title">Summary Builder</div>', unsafe_allow_html=True)
            builder_cols = st.columns(3)
            with builder_cols[0]:
                previous_summary_groups = st.session_state.get("atlas_export_summary_groups", [])
                if not isinstance(previous_summary_groups, list):
                    previous_summary_groups = []
                default_summary_groups = [column for column in ["ShipName", "ReportType"] if column in summary_builder_columns]
                valid_summary_group_defaults = [
                    column for column in previous_summary_groups
                    if column in summary_builder_columns
                ]
                if not valid_summary_group_defaults and "atlas_export_summary_groups" not in st.session_state:
                    valid_summary_group_defaults = default_summary_groups
                if valid_summary_group_defaults != previous_summary_groups:
                    st.session_state["atlas_export_summary_groups"] = valid_summary_group_defaults

                summary_group_fields = st.multiselect(
                    "Group by fields",
                    options=summary_builder_columns,
                    default=valid_summary_group_defaults,
                    key="atlas_export_summary_groups",
                    help="Choose the fields that define each summary row.",
                )
            with builder_cols[1]:
                previous_summary_values = st.session_state.get("atlas_export_summary_values", [])
                if not isinstance(previous_summary_values, list):
                    previous_summary_values = []
                valid_summary_value_defaults = [
                    column for column in previous_summary_values
                    if column in summary_value_options
                ]
                if valid_summary_value_defaults != previous_summary_values:
                    st.session_state["atlas_export_summary_values"] = valid_summary_value_defaults

                summary_value_fields = st.multiselect(
                    "Value fields",
                    options=summary_value_options,
                    default=valid_summary_value_defaults,
                    key="atlas_export_summary_values",
                    help="Choose one or more numeric columns to aggregate.",
                )
            with builder_cols[2]:
                summary_aggregation = st.selectbox(
                    "Aggregation",
                    options=["Average", "Sum", "Count", "Minimum", "Maximum", "Median"],
                    index=0,
                    key="atlas_export_summary_aggregation",
                )
        else:
            summary_group_fields = st.session_state.get("atlas_export_summary_groups", [])
            if not isinstance(summary_group_fields, list):
                summary_group_fields = []
            summary_value_fields = st.session_state.get("atlas_export_summary_values", [])
            if not isinstance(summary_value_fields, list):
                summary_value_fields = []
            summary_aggregation = st.session_state.get("atlas_export_summary_aggregation", "Average")

        summary_can_build = bool(summary_group_fields and summary_value_fields)
        if preview_mode == "Summary Analysis" and summary_can_build:
            displayed_table_df = build_summary_analysis(
                output_df,
                group_fields=summary_group_fields,
                value_fields=summary_value_fields,
                aggregation=summary_aggregation,
            )
            export_sheet_name = "Summary Analysis"
        elif preview_mode == "Summary Analysis":
            displayed_table_df = pd.DataFrame()
            export_sheet_name = "Summary Analysis"
            st.info("Select at least one Group by field and one Value field to preview Summary Analysis.")
        elif preview_mode == "Source Data":
            source_columns = [column for column in [*SOURCE_COLUMNS, "ParsedValue"] if column in filtered_long_for_options.columns]
            displayed_table_df = filtered_long_for_options[source_columns].copy()
            export_sheet_name = "Source Data"
        else:
            displayed_table_df = output_df.copy()
            export_sheet_name = "Clean Dataset"

        render_metric_cards(
            [
                ("Displayed Rows", f"{len(displayed_table_df):,}", "table_eye"),
                ("Selected Variables", f"{len(selected_variables):,}", "checked_columns"),
                ("Source Rows", f"{len(filtered_long_for_options):,}", "database_rows"),
                ("Available Variables", f"{len(variable_options):,}", "columns_plus"),
            ]
        )

        render_preview_table(displayed_table_df)
        if len(displayed_table_df) > TABLE_PREVIEW_ROW_LIMIT:
            st.caption(
                f"Showing first {TABLE_PREVIEW_ROW_LIMIT:,} of {len(displayed_table_df):,} rows. "
                "Excel export includes the full displayed table."
            )

        export_signature_payload = "|".join([
            preview_mode,
            ",".join(selected_vessels),
            selected_start.isoformat(),
            selected_end.isoformat(),
            ",".join(display_columns),
            ",".join(selected_variables),
            str(len(output_df)),
            str(len(displayed_table_df)),
            ",".join(summary_group_fields),
            ",".join(summary_value_fields),
            str(summary_aggregation),
            ",".join(displayed_table_df.columns.astype(str).tolist()) if not displayed_table_df.empty else "empty",
        ])
        current_export_signature = sha256(export_signature_payload.encode("utf-8")).hexdigest()

        export_ready = (
            st.session_state.get("atlas_export_signature") == current_export_signature
            and "atlas_export_bytes" in st.session_state
        )

        if st.button("Prepare displayed table Excel", type="primary", disabled=displayed_table_df.empty):
            with st.spinner("Preparing Excel file..."):
                st.session_state["atlas_export_bytes"] = to_displayed_table_excel_bytes(
                    displayed_table_df,
                    sheet_name=export_sheet_name,
                )
                st.session_state["atlas_summary_analysis_df"] = displayed_table_df if preview_mode == "Summary Analysis" else pd.DataFrame()
                st.session_state["atlas_export_signature"] = current_export_signature
                gc.collect()
            export_ready = True

        if export_ready:
            st.download_button(
                "Download displayed table Excel",
                data=st.session_state["atlas_export_bytes"],
                file_name="atlasflow_displayed_table.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.caption("Excel generation is prepared on demand. The download will contain only the visible table above.")

    with tab_reportpivots:
        reportpivots_output_df = render_wide_source_tab(
            "Noon & Manual Reports",
            reportpivots_df,
            reportpivots_metadata,
            "reportpivots",
            selected_vessels,
            selected_start,
            selected_end,
        )

    with tab_shippivots:
        shippivots_output_df = render_wide_source_tab(
            "15-Minute Operations",
            shippivots_df,
            shippivots_metadata,
            "shippivots",
            selected_vessels,
            selected_start,
            selected_end,
        )

    with tab_descriptive:
        render_descriptive_statistics_tab(
            output_df,
            reportpivots_output_df,
            shippivots_output_df,
        )

    with tab_export:
        st.markdown('<div class="section-title">AtlasFlow Export Center</div>', unsafe_allow_html=True)
        st.caption("Prepare a single workbook with Custom Analytics, Noon & Manual Reports, and 15-Minute Operations sheets from the current fleet/period selections.")
        render_metric_cards(
            [
                ("Custom Analytics Rows", f"{len(output_df):,}", "table_eye"),
                ("Noon & Manual Rows", f"{len(reportpivots_output_df):,}", "database_rows"),
                ("15-Minute Rows", f"{len(shippivots_output_df):,}", "checked_columns"),
            ]
        )

        multisource_signature_payload = "|".join([
            current_export_signature,
            str(len(reportpivots_output_df)),
            str(len(shippivots_output_df)),
            ",".join(reportpivots_output_df.columns.astype(str).tolist()) if not reportpivots_output_df.empty else "no_reportpivots",
            ",".join(shippivots_output_df.columns.astype(str).tolist()) if not shippivots_output_df.empty else "no_shippivots",
        ])
        multisource_signature = sha256(multisource_signature_payload.encode("utf-8")).hexdigest()
        multisource_ready = (
            st.session_state.get("atlas_multisource_export_signature") == multisource_signature
            and "atlas_multisource_export_bytes" in st.session_state
        )
        if st.button("Prepare full AtlasFlow workbook", type="primary"):
            with st.spinner("Preparing full AtlasFlow workbook..."):
                summary_analysis_df = pd.DataFrame()
                if summary_can_build:
                    summary_analysis_df = build_summary_analysis(
                        output_df,
                        group_fields=summary_group_fields,
                        value_fields=summary_value_fields,
                        aggregation=summary_aggregation,
                    )
                st.session_state["atlas_multisource_export_bytes"] = to_multisource_excel_bytes(
                    output_df,
                    summary_analysis_df if not summary_analysis_df.empty else None,
                    reportpivots_output_df,
                    shippivots_output_df,
                )
                st.session_state["atlas_multisource_export_signature"] = multisource_signature
                st.session_state["atlas_summary_analysis_df"] = summary_analysis_df
                st.session_state["atlas_export_signature"] = current_export_signature
            multisource_ready = True
        if multisource_ready:
            st.download_button(
                "Download full AtlasFlow workbook",
                data=st.session_state["atlas_multisource_export_bytes"],
                file_name="atlasflow_full_export.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.caption("The full workbook is prepared on demand to keep the live app responsive.")

    with tab_diagnostics:
        st.markdown('<div class="section-title">Diagnostics</div>', unsafe_allow_html=True)
        diagnostics = pd.DataFrame(
            {
                "Metric": [
                    "Selected vessels",
                    "API start date",
                    "API end date",
                    "Selected start",
                    "Selected end",
                    "API loaded at",
                    "API loaded local time",
                    "Loaded from snapshot",
                    "Snapshot saved at",
                    "Kept raw rows",
                    "Original API rows scanned",
                    "Discarded rows",
                    "Long rows prepared",
                    "Filtered long rows",
                    "Pivot rows before column filters",
                    "Pivot rows after column filters",
                    "Available variables",
                    "Selected variables",
                    "API pages",
                    "Downloaded MB",
                    "API fetch seconds",
                    "Prepare seconds",
                    "Hit API page limit",
                ],
                "Value": [
                    ", ".join(selected_vessels),
                    api_start_date.isoformat(),
                    api_end_date.isoformat(),
                    selected_start.isoformat(),
                    selected_end.isoformat(),
                    metadata.get("loaded_at_utc", "-"),
                    metadata.get("loaded_at_local", "-"),
                    str(metadata.get("loaded_from_snapshot", False)),
                    metadata.get("snapshot_saved_at_utc", "-"),
                    f"{metadata.get('kept_rows', metadata.get('rows', 0)):,}",
                    f"{metadata.get('scanned_rows', 0):,}",
                    f"{metadata.get('discarded_rows', 0):,}",
                    f"{len(long_df):,}",
                    f"{len(filtered_long_for_options):,}",
                    f"{len(pivot_df):,}",
                    f"{len(output_df):,}",
                    f"{len(variable_options):,}",
                    f"{len(selected_variables):,}",
                    f"{metadata.get('pages', 0):,}",
                    metadata.get("downloaded_mb", "-"),
                    metadata.get("fetch_seconds", "-"),
                    metadata.get("prepare_seconds", "-"),
                    str(metadata.get("hit_page_limit", "-")),
                ],
            }
        )
        st.dataframe(diagnostics, use_container_width=True, hide_index=True)

        with st.expander("First API URL", expanded=False):
            st.code(metadata.get("first_url", "-"), language="text")

        st.markdown('<div class="section-title">Available Variable Counts</div>', unsafe_allow_html=True)
        if st.button("Calculate variable counts"):
            value_counts = (
                filtered_long_for_options.get("ValueDescription", pd.Series(dtype="object"))
                .value_counts(dropna=False)
                .reset_index()
            )
            value_counts.columns = ["ValueDescription", "Rows"]
            st.dataframe(value_counts.head(500), use_container_width=True, hide_index=True)
        else:
            st.caption("Variable counts are calculated on demand so diagnostics do not slow normal loads.")



if __name__ == "__main__":
    main()
